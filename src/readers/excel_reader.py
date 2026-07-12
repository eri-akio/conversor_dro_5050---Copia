"""Leitor principal de arquivos Excel do Conversor DRO 5050.

Responsabilidades desta etapa:

- localizar e abrir um arquivo ``.xlsx``;
- verificar a presença das quatro abas obrigatórias;
- ler as quatro abas produtivas e fontes opcionais reconhecidas;
- ignorar outras abas adicionais sem tratá-las como erro;
- preservar o número original das linhas;
- preservar valores, fórmulas e metadados básicos das células;
- não normalizar nem corrigir dados.

As regras regulatórias, obrigatoriedades de campos, domínios e
conversões serão implementadas em etapas posteriores.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Any, Mapping
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from src.config import REQUIRED_SHEETS


class ExcelReaderError(Exception):
    """Erro conhecido durante a leitura estrutural do Excel."""

    def __init__(
        self,
        code: str,
        message: str,
        *,
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.details = dict(details or {})

    def __str__(self) -> str:
        return f"{self.code} — {self.message}"


@dataclass(frozen=True, slots=True)
class RawCell:
    """Célula lida sem normalização."""

    column_name: str
    coordinate: str
    value: Any
    data_type: str
    number_format: str
    is_formula: bool
    formula: str | None = None
    cached_value: Any = None

    @property
    def has_cached_formula_result(self) -> bool:
        return self.is_formula and self.cached_value is not None


@dataclass(frozen=True, slots=True)
class RawRow:
    """Linha de dados preservando o número original do Excel."""

    row_number: int
    cells: Mapping[str, RawCell]

    def get_cell(self, column_name: str) -> RawCell:
        """Retorna a célula correspondente à coluna solicitada."""

        try:
            return self.cells[column_name]
        except KeyError as error:
            raise KeyError(
                f"Coluna não encontrada na linha: {column_name}"
            ) from error

    def get_value(
        self,
        column_name: str,
        default: Any = None,
    ) -> Any:
        """Retorna somente o valor bruto da célula."""

        cell = self.cells.get(column_name)
        return default if cell is None else cell.value

    def as_values_dict(self) -> dict[str, Any]:
        """Converte a linha para um dicionário simples de valores."""

        return {
            column_name: cell.value
            for column_name, cell in self.cells.items()
        }


@dataclass(frozen=True, slots=True)
class RawSheet:
    """Conteúdo bruto de uma aba obrigatória."""

    name: str
    headers: tuple[str, ...]
    rows: tuple[RawRow, ...]
    ignored_empty_rows: int = 0

    @property
    def row_count(self) -> int:
        """Quantidade de linhas de dados efetivamente lidas."""

        return len(self.rows)

    @property
    def column_count(self) -> int:
        """Quantidade de colunas presentes no cabeçalho."""

        return len(self.headers)

    @property
    def formula_count(self) -> int:
        """Quantidade de células com fórmula na aba."""

        return sum(
            1
            for row in self.rows
            for cell in row.cells.values()
            if cell.is_formula
        )


@dataclass(frozen=True, slots=True)
class ExcelReadResult:
    """Resultado da leitura das abas produtivas e opcionais presentes."""

    source_path: Path
    file_size_bytes: int
    sheets: Mapping[str, RawSheet]
    additional_sheet_names: tuple[str, ...]

    def get_sheet(self, sheet_name: str) -> RawSheet:
        """Retorna uma aba lida pelo nome exato."""

        try:
            return self.sheets[sheet_name]
        except KeyError as error:
            raise KeyError(
                f"Aba não carregada: {sheet_name}"
            ) from error

    @property
    def total_rows(self) -> int:
        """Total de linhas de dados nas quatro abas."""

        return sum(sheet.row_count for sheet in self.sheets.values())

    @property
    def total_formulas(self) -> int:
        """Total de fórmulas encontradas nas quatro abas."""

        return sum(sheet.formula_count for sheet in self.sheets.values())


class ExcelWorkbookReader:
    """Serviço responsável pela leitura estrutural do arquivo Excel."""

    def __init__(
        self,
        required_sheets: tuple[str, ...] = REQUIRED_SHEETS,
    ) -> None:
        self.required_sheets = required_sheets

    def read(self, excel_path: str | Path) -> ExcelReadResult:
        """Abre o Excel e devolve as quatro abas como dados brutos."""

        source_path = self._validate_source_path(excel_path)
        workbook = None
        values_workbook = None

        try:
            workbook = load_workbook(
                filename=source_path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
            values_workbook = load_workbook(
                filename=source_path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )

            self._validate_required_sheets(
                workbook.sheetnames,
                source_path,
            )

            loaded_sheet_names = self.required_sheets
            sheets = {
                sheet_name: self._read_worksheet(
                    workbook[sheet_name],
                    values_workbook[sheet_name],
                )
                for sheet_name in loaded_sheet_names
            }

            additional_sheet_names = tuple(
                sheet_name
                for sheet_name in workbook.sheetnames
                if sheet_name not in loaded_sheet_names
            )

            return ExcelReadResult(
                source_path=source_path,
                file_size_bytes=source_path.stat().st_size,
                sheets=MappingProxyType(sheets),
                additional_sheet_names=additional_sheet_names,
            )

        except ExcelReaderError:
            raise
        except PermissionError as error:
            raise ExcelReaderError(
                "XLSX-READ-003",
                "Sem permissão para abrir o arquivo Excel.",
                details={"arquivo": str(source_path)},
            ) from error
        except (BadZipFile, InvalidFileException, KeyError) as error:
            raise ExcelReaderError(
                "XLSX-READ-004",
                "O arquivo não é um .xlsx válido ou está corrompido.",
                details={"arquivo": str(source_path)},
            ) from error
        except OSError as error:
            raise ExcelReaderError(
                "XLSX-READ-005",
                "Falha do sistema operacional ao ler o Excel.",
                details={
                    "arquivo": str(source_path),
                    "erro": str(error),
                },
            ) from error
        finally:
            if workbook is not None:
                workbook.close()
            if values_workbook is not None:
                values_workbook.close()

    @staticmethod
    def _validate_source_path(
        excel_path: str | Path,
    ) -> Path:
        """Valida existência, extensão e natureza do caminho."""

        raw_path = str(excel_path).strip()

        if not raw_path:
            raise ExcelReaderError(
                "XLSX-READ-001",
                "O caminho do arquivo Excel não foi informado.",
            )

        source_path = Path(raw_path).expanduser().resolve()

        if source_path.name.startswith("~$"):
            raise ExcelReaderError(
                "XLSX-READ-006",
                "Foi selecionado um arquivo temporário do Excel.",
                details={"arquivo": str(source_path)},
            )

        if source_path.suffix.lower() != ".xlsx":
            raise ExcelReaderError(
                "XLSX-READ-002",
                "Formato não suportado. Selecione um arquivo .xlsx.",
                details={"arquivo": str(source_path)},
            )

        if not source_path.exists():
            raise ExcelReaderError(
                "XLSX-READ-001",
                "O arquivo Excel não foi encontrado.",
                details={"arquivo": str(source_path)},
            )

        if not source_path.is_file():
            raise ExcelReaderError(
                "XLSX-READ-001",
                "O caminho informado não representa um arquivo.",
                details={"arquivo": str(source_path)},
            )

        return source_path

    def _validate_required_sheets(
        self,
        workbook_sheet_names: list[str],
        source_path: Path,
    ) -> None:
        """Confirma a presença exata das quatro abas obrigatórias."""

        missing = tuple(
            sheet_name
            for sheet_name in self.required_sheets
            if sheet_name not in workbook_sheet_names
        )

        if missing:
            raise ExcelReaderError(
                "XLSX-EST-001",
                "Uma ou mais abas obrigatórias não foram encontradas.",
                details={
                    "arquivo": str(source_path),
                    "abas_ausentes": missing,
                    "abas_encontradas": tuple(workbook_sheet_names),
                },
            )

    def _read_worksheet(
        self,
        worksheet: Worksheet,
        values_worksheet: Worksheet,
    ) -> RawSheet:
        """Lê cabeçalho e linhas de uma aba sem normalizar valores."""

        headers = self._read_headers(worksheet)

        rows: list[RawRow] = []
        ignored_empty_rows = 0

        formula_rows = worksheet.iter_rows(
            min_row=2,
            max_col=len(headers),
        )
        value_rows = values_worksheet.iter_rows(
            min_row=2,
            max_col=len(headers),
        )

        for excel_row_number, row_pair in enumerate(
            zip(formula_rows, value_rows, strict=True),
            start=2,
        ):
            excel_cells, value_cells = row_pair
            values = tuple(cell.value for cell in excel_cells)

            if self._is_completely_empty(values):
                ignored_empty_rows += 1
                continue

            row_cells: dict[str, RawCell] = {}

            for column_index, (header, cell, value_cell) in enumerate(
                zip(
                    headers,
                    excel_cells,
                    value_cells,
                    strict=True,
                ),
                start=1,
            ):
                coordinate = (
                    f"{get_column_letter(column_index)}"
                    f"{excel_row_number}"
                )
                data_type = str(
                    getattr(cell, "data_type", "") or ""
                )
                number_format = str(
                    getattr(cell, "number_format", "General")
                    or "General"
                )
                is_formula = data_type == "f"

                row_cells[header] = RawCell(
                    column_name=header,
                    coordinate=coordinate,
                    value=cell.value,
                    data_type=data_type,
                    number_format=number_format,
                    is_formula=is_formula,
                    formula=(
                        str(cell.value)
                        if is_formula
                        else None
                    ),
                    cached_value=(
                        value_cell.value
                        if is_formula
                        else None
                    ),
                )

            rows.append(
                RawRow(
                    row_number=excel_row_number,
                    cells=MappingProxyType(row_cells),
                )
            )

        return RawSheet(
            name=worksheet.title,
            headers=headers,
            rows=tuple(rows),
            ignored_empty_rows=ignored_empty_rows,
        )

    @staticmethod
    def _read_headers(
        worksheet: Worksheet,
    ) -> tuple[str, ...]:
        """Lê a primeira linha e rejeita cabeçalhos ambíguos."""

        header_row = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
            ),
            None,
        )

        if header_row is None:
            raise ExcelReaderError(
                "XLSX-EST-005",
                f"A aba {worksheet.title!r} não possui cabeçalho.",
                details={"aba": worksheet.title},
            )

        raw_headers = tuple(cell.value for cell in header_row)

        if not raw_headers or all(
            value is None for value in raw_headers
        ):
            raise ExcelReaderError(
                "XLSX-EST-005",
                f"A aba {worksheet.title!r} possui cabeçalho vazio.",
                details={"aba": worksheet.title},
            )

        headers: list[str] = []

        for column_index, raw_header in enumerate(
            raw_headers,
            start=1,
        ):
            if not isinstance(raw_header, str):
                raise ExcelReaderError(
                    "XLSX-EST-005",
                    (
                        f"O cabeçalho da aba {worksheet.title!r} "
                        "deve conter somente nomes textuais."
                    ),
                    details={
                        "aba": worksheet.title,
                        "coluna": column_index,
                        "valor": raw_header,
                    },
                )

            header = raw_header.strip()

            if not header:
                raise ExcelReaderError(
                    "XLSX-EST-005",
                    (
                        f"A aba {worksheet.title!r} possui uma "
                        "coluna sem nome."
                    ),
                    details={
                        "aba": worksheet.title,
                        "coluna": column_index,
                    },
                )

            headers.append(header)

        normalized_headers = [
            header.casefold()
            for header in headers
        ]
        duplicate_headers = sorted({
            headers[index]
            for index, normalized in enumerate(normalized_headers)
            if normalized_headers.count(normalized) > 1
        })

        if duplicate_headers:
            raise ExcelReaderError(
                "XLSX-EST-004",
                (
                    f"A aba {worksheet.title!r} possui cabeçalhos "
                    "duplicados."
                ),
                details={
                    "aba": worksheet.title,
                    "cabecalhos_duplicados": tuple(
                        duplicate_headers
                    ),
                },
            )

        return tuple(headers)

    @staticmethod
    def _is_completely_empty(
        values: tuple[Any, ...],
    ) -> bool:
        """Considera vazia somente a linha sem nenhum valor físico."""

        return all(value is None for value in values)


def read_excel(excel_path: str | Path) -> ExcelReadResult:
    """Atalho funcional para o leitor padrão do projeto."""

    return ExcelWorkbookReader().read(excel_path)

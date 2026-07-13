"""Geração do relatório XLSX usando ``openpyxl``."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.domain.reporting import (
    ExecutionReportData,
    FinalExecutionStatus,
)


TITLE_FILL = "17365D"
SECTION_FILL = "D9EAF7"
HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
BORDER_COLOR = "B7C9D6"

STATUS_FILLS = {
    FinalExecutionStatus.APT: "C6EFCE",
    FinalExecutionStatus.NOT_APT: "FFC7CE",
    FinalExecutionStatus.TECHNICAL_FAILURE: "D9D9D9",
}

OCCURRENCE_HEADERS = (
    "Etapa",
    "Linha",
    "idEvento",
    "Coluna",
    "Valor Original",
    "Valor Normalizado",
    "Regra",
    "Descrição da Regra",
    "Origem",
    "Gravidade",
    "Status",
    "Mensagem",
)


def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


class XlsxReportWriter:
    """Cria um relatório auditável e formatado em Excel.

    Saída simplificada:
    - Resumo
    - Ocorrencias
    """

    def write(
        self,
        data: ExecutionReportData,
        output_path: str | Path,
    ) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Resumo"
        workbook.create_sheet("Ocorrencias")

        self._build_occurrences(workbook, data)
        self._build_summary(workbook, data)
        workbook.save(path)

        if not path.is_file() or path.stat().st_size == 0:
            raise OSError(
                "O relatório XLSX não foi gravado corretamente."
            )

        return path

    def _build_summary(
        self,
        workbook: Workbook,
        data: ExecutionReportData,
    ) -> None:
        sheet = workbook["Resumo"]

        sheet.merge_cells("A1:H2")
        title = sheet["A1"]
        title.value = "RELATÓRIO DE EXECUÇÃO — DRO 5050"
        title.fill = _fill(TITLE_FILL)
        title.font = Font(bold=True, color=HEADER_FONT, size=16)
        title.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        for row in sheet["A1:H2"]:
            for cell in row:
                cell.fill = _fill(TITLE_FILL)

        metadata = (
            ("Status local", data.status_local.value),
            ("Status XSD", data.status_xsd.value),
            (
                "Validações externas ou históricas",
                data.status_dependencias.value,
            ),
            ("Resultado geral", data.general_status.value),
            ("Aptidão para envio", data.final_status.value),
        )

        self._section_title(sheet, "A3:B3", "Resultado da validação")

        bottom_side = Side(
            style="thin",
            color=BORDER_COLOR,
        )
        for row_index, (label, value) in enumerate(metadata, start=4):
            label_cell = sheet.cell(row=row_index, column=1, value=label)
            label_cell.fill = _fill(SECTION_FILL)
            label_cell.font = Font(bold=True)
            label_cell.border = Border(bottom=bottom_side)
            sheet.cell(row=row_index, column=2, value=value)

        status_row = 8
        sheet.cell(row=status_row, column=2).fill = _fill(
            STATUS_FILLS[data.final_status]
        )
        sheet.cell(row=status_row, column=2).font = Font(bold=True)

        self._section_title(sheet, "D3:E3", "INDICADORES DA EXECUÇÃO")
        metric_start = 4
        for row_index, (name, value) in enumerate(
            data.metrics,
            start=metric_start,
        ):
            name_cell = sheet.cell(row=row_index, column=4, value=name)
            name_cell.fill = _fill(SECTION_FILL)
            name_cell.font = Font(bold=True)
            sheet.cell(row=row_index, column=5, value=value)
        metric_end = max(metric_start, metric_start + len(data.metrics) - 1)

        self._section_title(sheet, "G3:H3", "CONTAGEM DAS OCORRÊNCIAS")
        labels_and_formulas = (
            (
                "ERRO IMPEDITIVO",
                data.severity_counts.get("ERRO IMPEDITIVO", 0),
            ),
            (
                "ERRO",
                data.severity_counts.get("ERRO", 0),
            ),
            (
                "AVISO",
                data.severity_counts.get("AVISO", 0),
            ),
            (
                "INFORMAÇÃO",
                data.severity_counts.get("INFORMAÇÃO", 0),
            ),
            (
                "REGRA NÃO EXECUTADA",
                sum(
                    record.status == "REGRA NÃO EXECUTADA"
                    for record in data.records
                ),
            ),
            (
                "REPROVADA",
                sum(
                    record.status == "REPROVADA"
                    for record in data.records
                ),
            ),
        )
        count_start = 4
        for row_index, (label, formula) in enumerate(
            labels_and_formulas,
            start=count_start,
        ):
            label_cell = sheet.cell(row=row_index, column=7, value=label)
            label_cell.fill = _fill(SECTION_FILL)
            label_cell.font = Font(bold=True)
            sheet.cell(row=row_index, column=8, value=formula)
        count_end = count_start + len(labels_and_formulas) - 1

        metadata_end = 3 + len(metadata)
        content_end = max(metric_end, count_end, metadata_end, 5)
        for row in sheet.iter_rows(
            min_row=4,
            max_row=content_end,
            min_col=1,
            max_col=8,
        ):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        widths = (24, 44, 4, 32, 18, 4, 26, 15)
        for column_index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        sheet.freeze_panes = "A3"

    @staticmethod
    def _section_title(sheet, cell_range: str, value: str) -> None:
        sheet.merge_cells(cell_range)
        cell = sheet[cell_range.split(":", maxsplit=1)[0]]
        cell.value = value
        cell.fill = _fill(HEADER_FILL)
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.alignment = Alignment(horizontal="center")
        for row in sheet[cell_range]:
            for merged_cell in row:
                merged_cell.fill = _fill(HEADER_FILL)

    def _build_occurrences(
        self,
        workbook: Workbook,
        data: ExecutionReportData,
    ) -> None:
        sheet = workbook["Ocorrencias"]
        sheet.append(OCCURRENCE_HEADERS)

        for record in data.records:
            sheet.append(
                (
                    record.stage,
                    record.row_numbers,
                    record.id_evento,
                    record.columns,
                    record.original_value,
                    record.normalized_value,
                    record.rule_code,
                    record.rule_description,
                    record.source,
                    record.severity,
                    record.status,
                    record.message,
                )
            )

        end_row = sheet.max_row
        end_col = len(OCCURRENCE_HEADERS)
        for cell in sheet[1]:
            cell.fill = _fill(HEADER_FILL)
            cell.font = Font(bold=True, color=HEADER_FONT)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        if data.records:
            table = Table(
                displayName="OccurrencesTable",
                ref=f"A1:{get_column_letter(end_col)}{end_row}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)

        sheet.freeze_panes = "A2"
        for row in sheet.iter_rows(
            min_row=2,
            max_row=end_row,
            min_col=1,
            max_col=end_col,
        ):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        widths = {
            "A": 24,
            "B": 12,
            "C": 20,
            "D": 24,
            "E": 30,
            "F": 30,
            "G": 18,
            "H": 42,
            "I": 30,
            "J": 22,
            "K": 22,
            "L": 48,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        if data.records:
            severity_range = f"J2:J{end_row}"
            sheet.conditional_formatting.add(
                severity_range,
                FormulaRule(
                    formula=['$J2="ERRO IMPEDITIVO"'],
                    fill=_fill("F4CCCC"),
                    font=Font(bold=True, color="9C0006"),
                ),
            )
            status_range = f"K2:K{end_row}"
            sheet.conditional_formatting.add(
                status_range,
                FormulaRule(
                    formula=['$K2="REPROVADA"'],
                    fill=_fill("F4CCCC"),
                    font=Font(bold=True, color="9C0006"),
                ),
            )
            sheet.conditional_formatting.add(
                status_range,
                FormulaRule(
                    formula=['$K2="REGRA NÃO EXECUTADA"'],
                    fill=_fill("E4DFEC"),
                    font=Font(color="5F497A"),
                ),
            )
            sheet.conditional_formatting.add(
                severity_range,
                FormulaRule(
                    formula=['$J2="ERRO"'],
                    fill=_fill("FCE5CD"),
                    font=Font(color="9C5700"),
                ),
            )
            sheet.conditional_formatting.add(
                severity_range,
                FormulaRule(
                    formula=['$J2="AVISO"'],
                    fill=_fill("FFF2CC"),
                    font=Font(color="7F6000"),
                ),
            )


def write_xlsx_report(
    data: ExecutionReportData,
    output_path: str | Path,
) -> Path:
    """Atalho funcional para o relatório Excel."""

    return XlsxReportWriter().write(data, output_path)

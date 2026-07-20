"""Testes das tabelas de sistemas e contas internas."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.config import (
    BASE_CREDIT_ACCOUNT_NAME_COLUMN,
    BASE_DEBIT_ACCOUNT_NAME_COLUMN,
    BASE_SOURCE_SYSTEM_NAME_COLUMN,
)
from src.domain.base_row_validation import RuleExecutionStatus
from src.mappers import group_base_rows
from src.readers import (
    read_and_normalize_base,
    read_excel,
    read_reference_tables,
)
from src.services.version_resolver import resolve_version
from src.validators import (
    validate_base_rows,
    validate_reference_tables,
)
from tests.test_event_grouping import (
    base_values,
    create_workbook,
)


def prepare_workbook(
    destination: Path,
    *,
    base_rows: list[dict[str, object]] | None = None,
    systems: list[tuple[object, object]] | None = None,
    accounts: list[tuple[object, object]] | None = None,
    system_headers: tuple[str, ...] = (
        "codigoSistema",
        "nomeSistema",
    ),
    account_headers: tuple[str, ...] = (
        "codigoConta",
        "nomeConta",
    ),
) -> None:
    rows = base_rows or [base_values()]
    create_workbook(destination, rows)

    workbook = load_workbook(destination)

    system_sheet = workbook["Sistemas_Origem"]
    system_sheet.delete_rows(1, system_sheet.max_row)
    system_sheet.append(list(system_headers))
    system_rows = (
        [("SIST0001", "Sistema Origem")]
        if systems is None
        else systems
    )
    for values in system_rows:
        system_sheet.append(list(values))

    account_sheet = workbook["Contas_Internas"]
    account_sheet.delete_rows(1, account_sheet.max_row)
    account_sheet.append(list(account_headers))
    account_rows = (
        [
            (
                "810000000000000000000001",
                "Conta Debito",
            ),
            (
                "410000000000000000000001",
                "Conta Credito",
            ),
        ]
        if accounts is None
        else accounts
    )
    for values in account_rows:
        account_sheet.append(list(values))

    workbook.save(destination)
    workbook.close()


def prepare_embedded_workbook(
    destination: Path,
    *,
    base_rows: list[dict[str, object]] | None = None,
    legacy_name_headers: bool = False,
) -> None:
    rows = base_rows or [base_values()]
    prepared_rows: list[dict[str, object]] = []
    for values in rows:
        prepared = dict(values)
        prepared.setdefault(
            BASE_SOURCE_SYSTEM_NAME_COLUMN,
            "Sistema Origem",
        )
        prepared.setdefault(
            BASE_DEBIT_ACCOUNT_NAME_COLUMN,
            "Conta Debito",
        )
        prepared.setdefault(
            BASE_CREDIT_ACCOUNT_NAME_COLUMN,
            "Conta Credito",
        )
        prepared_rows.append(prepared)

    create_workbook(destination, prepared_rows)
    workbook = load_workbook(destination)
    del workbook["Sistemas_Origem"]
    del workbook["Contas_Internas"]

    if legacy_name_headers:
        base = workbook["Base"]
        original_headers = tuple(
            cell.value for cell in base[1]
        )
        original_rows = tuple(
            dict(zip(
                original_headers,
                (cell.value for cell in row),
                strict=True,
            ))
            for row in base.iter_rows(
                min_row=2,
                max_row=base.max_row,
            )
        )
        ordered_headers: list[str] = []
        for header in original_headers:
            if header in {
                BASE_SOURCE_SYSTEM_NAME_COLUMN,
                BASE_DEBIT_ACCOUNT_NAME_COLUMN,
                BASE_CREDIT_ACCOUNT_NAME_COLUMN,
            }:
                continue
            ordered_headers.append(header)
            if header == "codSistemaOrigem":
                ordered_headers.append(
                    BASE_SOURCE_SYSTEM_NAME_COLUMN
                )
            elif header == "contaBalAnaliticoDebito":
                ordered_headers.append(
                    BASE_DEBIT_ACCOUNT_NAME_COLUMN
                )
            elif header == "contaBalAnaliticoCredito":
                ordered_headers.append(
                    BASE_CREDIT_ACCOUNT_NAME_COLUMN
                )

        workbook.remove(base)
        base = workbook.create_sheet("Base", 0)
        aliases = {
            BASE_SOURCE_SYSTEM_NAME_COLUMN: "nomeSistema",
            BASE_DEBIT_ACCOUNT_NAME_COLUMN: "nomeConta",
            BASE_CREDIT_ACCOUNT_NAME_COLUMN: "nomeConta",
        }
        base.append([
            aliases.get(header, header)
            for header in ordered_headers
        ])
        for values in original_rows:
            base.append([
                values.get(header)
                for header in ordered_headers
            ])

    workbook.save(destination)
    workbook.close()


def process_reference_tables(path: Path):
    excel = read_excel(path)
    profile = resolve_version("2026-06").profile
    assert profile is not None

    normalization = read_and_normalize_base(
        excel,
        profile,
    )
    row_validation = validate_base_rows(
        normalization,
        profile,
    )
    grouping = group_base_rows(
        normalization,
        row_validation,
    )
    read_result = read_reference_tables(excel)
    validation = validate_reference_tables(
        read_result,
        grouping,
    )
    return read_result, validation


def find_rule(result, code: str):
    matches = [
        item
        for item in result.rule_results
        if item.code == code
    ]
    assert matches
    return matches


def test_valid_tables_and_references_pass(
    tmp_path: Path,
) -> None:
    path = tmp_path / "valid.xlsx"
    prepare_workbook(path)

    read_result, validation = process_reference_tables(path)

    assert read_result.systems.row_count == 1
    assert read_result.accounts.row_count == 2
    assert read_result.is_valid
    assert validation.is_valid
    assert validation.is_fully_verified

    assert find_rule(
        validation,
        "DRO001102",
    )[0].status == RuleExecutionStatus.PASSED
    assert find_rule(
        validation,
        "DRO001101",
    )[0].status == RuleExecutionStatus.PASSED
    assert find_rule(
        validation,
        "DRO001321",
    )[0].status == RuleExecutionStatus.PASSED
    assert find_rule(
        validation,
        "DRO001401",
    )[0].status == RuleExecutionStatus.PASSED
    assert find_rule(
        validation,
        "DRO001402",
    )[0].status == RuleExecutionStatus.PASSED


def test_embedded_references_are_deduplicated_and_validated(
    tmp_path: Path,
) -> None:
    path = tmp_path / "embedded.xlsx"
    first = base_values()
    second = dict(first)
    second["idEvento"] = "EVT0002"
    second["codigoEventoOrigem"] = "ORIG0002"
    prepare_embedded_workbook(
        path,
        base_rows=[first, second],
    )

    read_result, validation = process_reference_tables(path)

    assert read_result.systems.sheet_name == "Base"
    assert read_result.accounts.sheet_name == "Base"
    assert read_result.systems.row_count == 1
    assert read_result.accounts.row_count == 2
    assert validation.is_valid
    assert validation.is_fully_verified


def test_embedded_legacy_name_headers_are_canonicalized(
    tmp_path: Path,
) -> None:
    path = tmp_path / "embedded_aliases.xlsx"
    prepare_embedded_workbook(
        path,
        legacy_name_headers=True,
    )

    read_result, validation = process_reference_tables(path)

    assert read_result.systems.row_count == 1
    assert read_result.accounts.row_count == 2
    assert validation.is_valid


def test_embedded_system_code_with_different_names_fails(
    tmp_path: Path,
) -> None:
    path = tmp_path / "embedded_conflict.xlsx"
    first = base_values()
    second = dict(first)
    second["idEvento"] = "EVT0002"
    second["codigoEventoOrigem"] = "ORIG0002"
    first[BASE_SOURCE_SYSTEM_NAME_COLUMN] = "Sistema Um"
    second[BASE_SOURCE_SYSTEM_NAME_COLUMN] = "Sistema Dois"
    prepare_embedded_workbook(
        path,
        base_rows=[first, second],
    )

    _, validation = process_reference_tables(path)
    rule = find_rule(validation, "DRO001102")[0]

    assert rule.status == RuleExecutionStatus.FAILED
    assert rule.row_numbers == (2, 3)
    assert not validation.is_valid


def test_missing_required_column_blocks_table(
    tmp_path: Path,
) -> None:
    path = tmp_path / "missing_column.xlsx"
    prepare_workbook(
        path,
        system_headers=("codigoSistema",),
        systems=[("SIST0001",)],
    )

    read_result, validation = process_reference_tables(path)

    assert not read_result.systems.is_valid
    assert not validation.is_valid
    assert read_result.systems.missing_columns == (
        "nomeSistema",
    )
    assert any(
        item.code == "TBL-SIS-EST-001"
        and item.status == RuleExecutionStatus.FAILED
        for item in validation.rule_results
    )


def test_empty_table_blocks_processing(
    tmp_path: Path,
) -> None:
    path = tmp_path / "empty.xlsx"
    prepare_workbook(
        path,
        systems=[],
    )

    read_result, validation = process_reference_tables(path)

    assert read_result.systems.row_count == 0
    assert not validation.is_valid
    assert any(
        item.code == "TBL-SIS-EST-002"
        for item in validation.failed_rules
    )


def test_duplicate_system_code_fails_dro001102(
    tmp_path: Path,
) -> None:
    path = tmp_path / "duplicate_system.xlsx"
    prepare_workbook(
        path,
        systems=[
            ("SIST0001", "Sistema Um"),
            ("SIST0001", "Sistema Dois"),
        ],
    )

    _, validation = process_reference_tables(path)
    rule = find_rule(validation, "DRO001102")[0]

    assert rule.status == RuleExecutionStatus.FAILED
    assert rule.row_numbers == (2, 3)
    assert not validation.is_valid


def test_duplicate_account_code_fails_dro001101(
    tmp_path: Path,
) -> None:
    path = tmp_path / "duplicate_account.xlsx"
    prepare_workbook(
        path,
        accounts=[
            (
                "810000000000000000000001",
                "Conta Um",
            ),
            (
                "810000000000000000000001",
                "Conta Dois",
            ),
            (
                "410000000000000000000001",
                "Conta Credito",
            ),
        ],
    )

    _, validation = process_reference_tables(path)
    rule = find_rule(validation, "DRO001101")[0]

    assert rule.status == RuleExecutionStatus.FAILED
    assert not validation.is_valid


def test_missing_system_reference_fails_dro001321(
    tmp_path: Path,
) -> None:
    path = tmp_path / "missing_system_reference.xlsx"
    prepare_workbook(
        path,
        systems=[("OUTRO001", "Outro Sistema")],
    )

    _, validation = process_reference_tables(path)
    rule = find_rule(validation, "DRO001321")[0]

    assert rule.status == RuleExecutionStatus.FAILED
    assert "SIST0001" in rule.message
    assert not validation.is_valid


def test_missing_account_references_fail_official_rules(
    tmp_path: Path,
) -> None:
    path = tmp_path / "missing_accounts.xlsx"
    prepare_workbook(
        path,
        accounts=[("999999", "Conta Outra")],
    )

    _, validation = process_reference_tables(path)

    assert find_rule(
        validation,
        "DRO001401",
    )[0].status == RuleExecutionStatus.FAILED
    assert find_rule(
        validation,
        "DRO001402",
    )[0].status == RuleExecutionStatus.FAILED
    assert not validation.is_valid


def test_duplicate_referenced_code_is_not_treated_as_valid(
    tmp_path: Path,
) -> None:
    path = tmp_path / "ambiguous_system.xlsx"
    prepare_workbook(
        path,
        systems=[
            ("SIST0001", "Sistema Um"),
            ("SIST0001", "Sistema Dois"),
        ],
    )

    _, validation = process_reference_tables(path)
    usage_rule = find_rule(validation, "DRO001321")[0]

    assert usage_rule.status == RuleExecutionStatus.NOT_EXECUTED
    assert not validation.is_fully_verified


def test_short_system_code_is_accepted_by_xsd_precedence(
    tmp_path: Path,
) -> None:
    values = base_values()
    values["codSistemaOrigem"] = "A"

    path = tmp_path / "short_code.xlsx"
    prepare_workbook(
        path,
        base_rows=[values],
        systems=[("A", "Sistema A")],
    )

    read_result, validation = process_reference_tables(path)

    assert read_result.systems.records[0].code == "A"
    assert validation.is_valid


def test_accented_name_is_rejected_by_xsd_pattern(
    tmp_path: Path,
) -> None:
    path = tmp_path / "accented_name.xlsx"
    prepare_workbook(
        path,
        systems=[("SIST0001", "Sistema Jurídico")],
    )

    read_result, validation = process_reference_tables(path)
    record = read_result.systems.records[0]

    assert record.name_result.is_invalid
    assert record.name_result.issue_code == "TBL-NOME-FMT-001"
    assert not validation.is_valid


def test_numeric_account_code_is_rejected_to_preserve_zeros(
    tmp_path: Path,
) -> None:
    path = tmp_path / "numeric_account.xlsx"
    prepare_workbook(
        path,
        accounts=[
            (810000000000000000000001, "Conta Debito"),
            (
                "410000000000000000000001",
                "Conta Credito",
            ),
        ],
    )

    read_result, validation = process_reference_tables(path)

    assert read_result.accounts.records[0].code_result.is_invalid
    assert not validation.is_valid


def test_formula_in_reference_table_is_invalid(
    tmp_path: Path,
) -> None:
    path = tmp_path / "formula.xlsx"
    prepare_workbook(path)

    workbook = load_workbook(path)
    workbook["Sistemas_Origem"]["A2"] = '=CONCAT("SIST","0001")'
    workbook.save(path)
    workbook.close()

    read_result, validation = process_reference_tables(path)

    assert (
        read_result.systems.records[0].code_result.issue_code
        == "TBL-FORMULA-001"
    )
    assert not validation.is_valid


def test_unused_codes_are_informational(
    tmp_path: Path,
) -> None:
    path = tmp_path / "unused.xlsx"
    prepare_workbook(
        path,
        systems=[
            ("SIST0001", "Sistema Origem"),
            ("NAOUSADO", "Sistema Nao Usado"),
        ],
        accounts=[
            (
                "810000000000000000000001",
                "Conta Debito",
            ),
            (
                "410000000000000000000001",
                "Conta Credito",
            ),
            ("123", "Conta Nao Usada"),
        ],
    )

    _, validation = process_reference_tables(path)

    assert validation.is_valid
    assert validation.unused_system_codes == ("NAOUSADO",)
    assert validation.unused_account_codes == ("123",)


def test_sample_workbook_passes_reference_validation() -> None:
    path = (
        Path(__file__).parent
        / "fixtures"
        / "DRO_5050_planilha_testes.xlsx"
    )

    read_result, validation = process_reference_tables(path)

    assert read_result.systems.row_count == 5
    assert read_result.accounts.row_count == 10
    assert validation.is_valid
    assert validation.is_fully_verified
    assert validation.unused_system_codes == ()
    assert validation.unused_account_codes == ()

    assert len(find_rule(validation, "DRO001321")) == 15
    assert len(find_rule(validation, "DRO001401")) == 30
    assert len(find_rule(validation, "DRO001402")) == 30
    assert all(
        item.status == RuleExecutionStatus.PASSED
        for item in validation.rule_results
    )

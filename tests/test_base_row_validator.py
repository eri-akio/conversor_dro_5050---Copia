"""Testes das obrigatoriedades e relações locais da aba Base."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.config import BASE_ALL_COLUMNS
from src.domain.base_row_validation import (
    BaseRowKind,
    RuleExecutionStatus,
)
from src.readers import read_and_normalize_base, read_excel
from src.services.version_resolver import resolve_version
from src.validators.base_row_validator import validate_base_rows


def valid_row_values() -> dict[str, object]:
    """Linha válida para as regras locais do perfil 2026-06."""

    return {
        "Source.Name": "base.xlsx",
        "idEvento": "EVT0001",
        "categoriaNivel1": "3",
        "categoriaNivel2": "31",
        "tipoAvaliacao": "I",
        "unidadeNegocio": "1",
        "dataDescoberta": "2025-06-10",
        "dataOcorrencia": "2025-06-08",
        "totalPerdaEfetiva": "1500,00",
        "totalProvisao": "500,00",
        "totalRecuperado": "0,00",
        "valorTotalRisco": "12.000.000,00",
        "naturezaContingencia": "TRA",
        "codSistemaOrigem": "SIST0001",
        "codigoEventoOrigem": "ORIG0001",
        "descricaoEvento": "Evento de teste local.",
        "riscoAssociado": "NA",
        "ligacaoRiscoSocioambiental": "N",
        "ligadoRiscoCibernetico": "N",
        "negocioDescontinuado": "N",
        "idBacen": "Z1234567",
        "probabilidadePerda": "PR",
        "valorRisco": "11.999.500,00",
        "dataContabilizacao": "2025-06-12",
        "contaBalAnaliticoDebito": "810000000000000000000001",
        "contaBalAnaliticoCredito": "410000000000000000000001",
        "contaCosifDebito": "10000007",
        "contaCosifCredito": "20000006",
        "valorPerdaEfetiva": "1500,00",
        "valorProvisao": "500,00",
        "valorRecuperacao": "0,00",
        "fonteRecuperacao": "NA",
        "idEventoAgregador": None,
        "dataExclusao": None,
        "motivoExclusao": None,
    }


def create_workbook(
    destination: Path,
    *,
    rows: list[dict[str, object]],
    data_base: str,
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    base = workbook.create_sheet("Base")
    base.append(list(BASE_ALL_COLUMNS))
    for row_values in rows:
        base.append(
            [
                row_values.get(column_name)
                for column_name in BASE_ALL_COLUMNS
            ]
        )

    header = workbook.create_sheet("Cabecalho")
    header.append(
        [
            "codigoDocumento",
            "dataBase",
            "codigoConglomerado",
            "cnpj",
            "tipoRemessa",
            "opcaoPorProvisaoAcumulada",
        ]
    )
    header.append(
        [
            "5050",
            data_base,
            "C0099999",
            "99999999",
            "I",
            "N",
        ]
    )

    systems = workbook.create_sheet("Sistemas_Origem")
    systems.append(["codigoSistema", "nomeSistema"])
    systems.append(["SIST0001", "Sistema de teste"])

    accounts = workbook.create_sheet("Contas_Internas")
    accounts.append(["codigoConta", "nomeConta"])
    accounts.append(
        ["810000000000000000000001", "Conta de débito"]
    )
    accounts.append(
        ["410000000000000000000001", "Conta de crédito"]
    )

    workbook.save(destination)
    workbook.close()


def validate_rows(
    tmp_path: Path,
    rows: list[dict[str, object]],
    *,
    data_base: str = "2026-06",
):
    excel_path = tmp_path / "entrada.xlsx"
    create_workbook(
        excel_path,
        rows=rows,
        data_base=data_base,
    )

    selection = resolve_version(data_base)
    assert selection.profile is not None

    normalization = read_and_normalize_base(
        read_excel(excel_path),
        selection.profile,
    )
    return validate_base_rows(
        normalization,
        selection.profile,
    )


def result_for_code(row_result, code: str):
    matches = [
        result
        for result in row_result.rule_results
        if result.code == code
    ]
    assert len(matches) == 1
    return matches[0]


def test_valid_row_passes_local_rules_but_keeps_deferred_rules(
    tmp_path: Path,
) -> None:
    result = validate_rows(tmp_path, [valid_row_values()])

    assert result.is_locally_valid
    assert not result.is_fully_verified
    assert result.locally_valid_row_count == 1
    assert result.invalid_row_count == 0

    row = result.rows[0]
    assert row.row_kind == BaseRowKind.INDIVIDUALIZED
    assert result_for_code(
        row,
        "DRO001231",
    ).status == RuleExecutionStatus.PASSED
    assert result_for_code(
        row,
        "DRO001312",
    ).status == RuleExecutionStatus.DEFERRED
    assert result_for_code(
        row,
        "DRO001241",
    ).status == RuleExecutionStatus.NOT_EXECUTED


def test_sample_workbook_has_no_local_failures() -> None:
    sample_path = (
        Path(__file__).parent
        / "fixtures"
        / "DRO_5050_planilha_testes.xlsx"
    )
    selection = resolve_version("2026-06")
    assert selection.profile is not None

    normalization = read_and_normalize_base(
        read_excel(sample_path),
        selection.profile,
    )
    result = validate_base_rows(
        normalization,
        selection.profile,
    )

    assert result.row_count == 30
    assert result.is_locally_valid
    assert result.invalid_row_count == 0
    assert not result.is_fully_verified


def test_missing_required_event_field_is_rejected(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["idEvento"] = None

    result = validate_rows(tmp_path, [values])
    row = result.rows[0]
    rule = result_for_code(row, "BASE-XSD-REQ-001")

    assert not result.is_locally_valid
    assert rule.status == RuleExecutionStatus.FAILED
    assert "idEvento" in rule.columns


def test_date_and_category_relations_are_validated(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["categoriaNivel1"] = "3"
    values["categoriaNivel2"] = "41"
    values["dataDescoberta"] = "2025-06-01"
    values["dataOcorrencia"] = "2025-06-08"

    result = validate_rows(tmp_path, [values])
    row = result.rows[0]

    assert result_for_code(
        row,
        "BASE-REL-CAT-001",
    ).status == RuleExecutionStatus.FAILED
    assert result_for_code(
        row,
        "DRO001201",
    ).status == RuleExecutionStatus.FAILED


def test_fields_required_after_2021_are_checked(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    for field_name in (
        "dataDescoberta",
        "categoriaNivel2",
        "riscoAssociado",
        "ligacaoRiscoSocioambiental",
        "ligadoRiscoCibernetico",
    ):
        values[field_name] = None

    result = validate_rows(tmp_path, [values])
    row = result.rows[0]

    for code in (
        "DRO001202",
        "DRO001212",
        "DRO001251",
        "DRO001252",
        "DRO001253",
    ):
        assert result_for_code(
            row,
            code,
        ).status == RuleExecutionStatus.FAILED


def test_description_uses_instruction_precedence(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["totalPerdaEfetiva"] = "600.000,00"
    values["totalProvisao"] = "400.000,00"
    values["descricaoEvento"] = None

    result = validate_rows(tmp_path, [values])
    row = result.rows[0]

    assert result_for_code(
        row,
        "BASE-OBR-DESC-001",
    ).status == RuleExecutionStatus.FAILED
    assert result_for_code(
        row,
        "DRO001241",
    ).status == RuleExecutionStatus.NOT_EXECUTED


def test_provision_is_incompatible_with_assessment_na(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["tipoAvaliacao"] = "NA"
    values["naturezaContingencia"] = "NA"
    values["totalProvisao"] = "100,00"
    values["valorProvisao"] = "100,00"
    values["valorTotalRisco"] = None
    values["probabilidadePerda"] = None
    values["valorRisco"] = None

    result = validate_rows(tmp_path, [values])
    row = result.rows[0]

    assert result_for_code(
        row,
        "DRO001301",
    ).status == RuleExecutionStatus.FAILED


def test_probability_pair_and_massified_rule(
    tmp_path: Path,
) -> None:
    pair_values = valid_row_values()
    pair_values["valorRisco"] = None

    massified_values = valid_row_values()
    massified_values["idEvento"] = "EVT0002"
    massified_values["tipoAvaliacao"] = "M"

    result = validate_rows(
        tmp_path,
        [pair_values, massified_values],
    )

    assert result_for_code(
        result.rows[0],
        "BASE-PROB-001",
    ).status == RuleExecutionStatus.FAILED
    assert result_for_code(
        result.rows[1],
        "DRO001313",
    ).status == RuleExecutionStatus.FAILED


def test_accounting_requirements_and_account_pairs(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["contaCosifDebito"] = None
    values["contaBalAnaliticoCredito"] = None
    values["contaCosifCredito"] = None

    result = validate_rows(tmp_path, [values])
    row = result.rows[0]

    assert result_for_code(
        row,
        "DRO001441",
    ).status == RuleExecutionStatus.FAILED
    assert result_for_code(
        row,
        "DRO001451",
    ).status == RuleExecutionStatus.FAILED


def test_recovery_sign_and_source_are_validated(
    tmp_path: Path,
) -> None:
    positive = valid_row_values()
    positive["valorRecuperacao"] = "100,00"

    without_source = valid_row_values()
    without_source["idEvento"] = "EVT0002"
    without_source["valorRecuperacao"] = "-100,00"
    without_source["fonteRecuperacao"] = "NA"

    result = validate_rows(
        tmp_path,
        [positive, without_source],
    )

    assert result_for_code(
        result.rows[0],
        "DRO001411",
    ).status == RuleExecutionStatus.FAILED
    assert result_for_code(
        result.rows[1],
        "DRO001421",
    ).status == RuleExecutionStatus.FAILED


def test_future_recovery_must_be_exclusive(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["valorRecuperacao"] = "-100,00"
    values["fonteRecuperacao"] = "S"
    values["valorPerdaEfetiva"] = "50,00"
    values["valorProvisao"] = "0,00"

    result = validate_rows(
        tmp_path,
        [values],
        data_base="2026-12",
    )

    assert result_for_code(
        result.rows[0],
        "BASE-REC-EXCL-001",
    ).status == RuleExecutionStatus.FAILED


def test_future_excluded_row_has_its_own_required_fields(
    tmp_path: Path,
) -> None:
    values = {
        column_name: None
        for column_name in BASE_ALL_COLUMNS
    }
    values.update(
        {
            "idEvento": "EVTEXC01",
            "dataExclusao": "2026-12-31",
            "motivoExclusao": "8 - Outros",
        }
    )

    result = validate_rows(
        tmp_path,
        [values],
        data_base="2026-12",
    )
    row = result.rows[0]

    assert row.row_kind == BaseRowKind.EXCLUDED
    assert row.is_locally_valid
    assert not row.is_fully_verified
    assert result_for_code(
        row,
        "BASE-EXCL-001",
    ).status == RuleExecutionStatus.PASSED
    assert result_for_code(
        row,
        "BASE-EXCL-DOM-001",
    ).status == RuleExecutionStatus.NOT_EXECUTED


def test_future_excluded_row_requires_reason_and_date(
    tmp_path: Path,
) -> None:
    values = {
        column_name: None
        for column_name in BASE_ALL_COLUMNS
    }
    values.update(
        {
            "idEvento": "EVTEXC01",
            "dataExclusao": "2026-12-31",
            "motivoExclusao": None,
        }
    )

    result = validate_rows(
        tmp_path,
        [values],
        data_base="2026-12",
    )
    row = result.rows[0]

    assert not row.is_locally_valid
    assert result_for_code(
        row,
        "BASE-EXCL-001",
    ).status == RuleExecutionStatus.FAILED


def test_total_recovery_limit_and_signs(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["totalPerdaEfetiva"] = "1.000,00"
    values["totalProvisao"] = "0,00"
    values["totalRecuperado"] = "-2.000,00"

    result = validate_rows(tmp_path, [values])
    row = result.rows[0]

    assert result_for_code(
        row,
        "DRO001232",
    ).status == RuleExecutionStatus.FAILED

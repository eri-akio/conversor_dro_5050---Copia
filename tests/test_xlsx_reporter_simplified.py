"""Testes da estrutura simplificada do relatório XLSX."""

from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from src.domain.reporting import (
    DependentValidationStatus,
    ExternalValidationStatus,
    ExecutionReportData,
    FinalExecutionStatus,
    GeneralValidationStatus,
    HistoricalValidationStatus,
    LocalValidationStatus,
    ReportRecord,
    XsdValidationSummaryStatus,
)
from src.reporters.xlsx_reporter import (
    OCCURRENCE_HEADERS,
    XlsxReportWriter,
)


REMOVED_SUMMARY_LABELS = {
    "Duração (segundos)",
    "Execução",
    "Arquivo de entrada",
    "dataBase",
    "Arquivo XML",
    "Início",
    "Fim",
    "Perfil",
    "XSD selecionado",
    "Mensagem final",
}

REMOVED_SUMMARY_NOTE = (
    "O resultado APTO exige XML válido no XSD, ausência de erros impeditivos "
    "locais e nenhuma regra regulatória pendente."
)

REMOVED_OCCURRENCE_HEADERS = {
    "Execução",
    "Data/Hora",
    "Arquivo de Entrada",
    "Arquivo XML",
    "Aba",
    "Dependência",
    "Resultado Final",
    "Versão",
    "Sugestão",
    "Escopo",
    "Resultado Definitivo",
}

EXPECTED_OCCURRENCE_HEADERS = (
    "Etapa",
    "Linha",
    "idEvento",
    "Coluna",
    "Valor Original",
    "Valor Normalizado",
    "Regra",
    "Descrição da Regra",
    "Origem",
    "Gravidade",
    "Status",
    "Mensagem",
)


class SimplifiedXlsxReporterTests(unittest.TestCase):
    def test_report_contains_only_requested_summary_and_columns(self) -> None:
        now = datetime.now().astimezone()
        record = ReportRecord(
            execution_id="exec-1",
            executed_at=now,
            final_result=FinalExecutionStatus.NOT_APT,
            input_file="entrada.xlsx",
            xml_file="saida.xml",
            stage="VALIDAÇÃO",
            sheet_name="Base",
            row_numbers="2",
            id_evento="EVT-1",
            columns="A",
            original_value="x",
            normalized_value="X",
            rule_code="REG-1",
            rule_description="Regra de teste",
            source="Teste",
            version="1",
            severity="ERRO IMPEDITIVO",
            status="REPROVADA",
            suggestion="Corrigir",
            message="Valor inválido",
        )
        data = ExecutionReportData(
            execution_id="exec-1",
            started_at=now,
            finished_at=now + timedelta(seconds=2),
            input_path=Path("entrada.xlsx"),
            xml_path=Path("saida.xml"),
            xsd_path=Path("schema.xsd"),
            data_base="2026-06",
            profile_code="DRO_2025_06",
            status_local=LocalValidationStatus.APPROVED,
            status_xsd=XsdValidationSummaryStatus.APPROVED,
            status_externo=ExternalValidationStatus.PENDING,
            status_historico=(
                HistoricalValidationStatus.PENDING
            ),
            status_dependencias=DependentValidationStatus.PENDING,
            general_status=GeneralValidationStatus.PENDING,
            final_status=FinalExecutionStatus.NOT_APT,
            final_message="Há pendências",
            records=(record,),
            pre_rules=(),
            post_rules=(),
            metrics=(("Registros", 1),),
        )

        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "relatorio.xlsx"
            XlsxReportWriter().write(data, output)
            workbook = load_workbook(output, data_only=False)

        self.assertEqual(workbook.sheetnames, ["Resumo", "Ocorrencias"])

        summary = workbook["Resumo"]
        summary_values = {
            cell.value
            for row in summary.iter_rows()
            for cell in row
            if isinstance(cell.value, str)
        }
        self.assertTrue(REMOVED_SUMMARY_LABELS.isdisjoint(summary_values))
        self.assertNotIn(REMOVED_SUMMARY_NOTE, summary_values)
        self.assertNotIn("Há pendências", summary_values)
        self.assertEqual(summary["A3"].value, "Resultado da validação")
        self.assertIn("A3:B3", {str(item) for item in summary.merged_cells.ranges})
        self.assertEqual(summary["A3"]._style, summary["D3"]._style)
        self.assertEqual(summary["A4"].value, "Status local")
        self.assertEqual(summary["B4"].value, "APROVADO")
        self.assertEqual(summary["A5"].value, "Status XSD")
        self.assertEqual(summary["B5"].value, "APROVADO")
        self.assertEqual(
            summary["A6"].value,
            "Validações externas ou históricas",
        )
        self.assertEqual(summary["B6"].value, "PENDENTE")
        self.assertEqual(summary["A7"].value, "Resultado geral")
        self.assertEqual(summary["B7"].value, "PENDENTE")
        self.assertEqual(summary["A8"].value, "Aptidão para envio")
        self.assertEqual(summary["B8"].value, "NÃO APTO PARA ENVIO")
        self.assertIsNone(summary["A9"].value)
        self.assertIsNone(summary["B9"].value)
        self.assertEqual(summary["H4"].value, 1)
        self.assertIsInstance(summary["H4"].value, int)
        self.assertEqual(summary["H8"].value, 0)
        self.assertEqual(summary["H9"].value, 1)

        occurrences = workbook["Ocorrencias"]
        headers = tuple(cell.value for cell in occurrences[1])
        self.assertEqual(headers, OCCURRENCE_HEADERS)
        self.assertEqual(headers, EXPECTED_OCCURRENCE_HEADERS)
        self.assertEqual(len(headers), 12)
        self.assertTrue(REMOVED_OCCURRENCE_HEADERS.isdisjoint(headers))
        self.assertEqual(occurrences["A2"].value, "VALIDAÇÃO")
        self.assertEqual(occurrences["J2"].value, "ERRO IMPEDITIVO")
        self.assertEqual(occurrences["K2"].value, "REPROVADA")
        self.assertEqual(occurrences["L2"].value, "Valor inválido")
        self.assertEqual(occurrences.freeze_panes, "A2")
        self.assertEqual(occurrences.tables["OccurrencesTable"].ref, "A1:L2")


if __name__ == "__main__":
    unittest.main()

"""Testes da validação estrutural da aba Base."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.config import (
    BASE_ALL_COLUMNS,
    BASE_CONFIRMED_REQUIRED_COLUMNS,
    BASE_EMBEDDED_REFERENCE_COLUMNS,
    BASE_FUTURE_COLUMNS,
)
from src.readers.excel_reader import read_excel
from src.services.version_resolver import resolve_version
from src.validators.base_structure_validator import (
    validate_base_structure,
)


def create_workbook(
    destination: Path,
    *,
    base_columns: tuple[str, ...],
    include_base_row: bool = True,
    include_reference_sheets: bool = True,
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    base = workbook.create_sheet("Base")
    base.append(list(base_columns))

    if include_base_row:
        base.append(
            [
                f"valor-{index}"
                for index, _ in enumerate(
                    base_columns,
                    start=1,
                )
            ]
        )

    header = workbook.create_sheet("Cabecalho")
    header.append(
        [
            "codigoDocumento",
            "dataBase",
            "codigoConglomerado",
            "cnpj",
            "tipoRemessa",
            "opcaoPorProvisaoAcumulada",
        ]
    )
    header.append(
        [
            "5050",
            "2026-06",
            "C0099999",
            "99999999",
            "I",
            "N",
        ]
    )

    if include_reference_sheets:
        systems = workbook.create_sheet("Sistemas_Origem")
        systems.append(["codigoSistema", "nomeSistema"])
        systems.append(["SIS001", "Sistema"])

        accounts = workbook.create_sheet("Contas_Internas")
        accounts.append(["codigoConta", "nomeConta"])
        accounts.append(["0001", "Conta"])

    workbook.save(destination)
    workbook.close()


def get_profile(data_base: str):
    selection = resolve_version(data_base)
    assert selection.profile is not None
    return selection.profile


def test_legacy_profile_accepts_confirmed_columns(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "legado.xlsx"
    create_workbook(
        excel_path,
        base_columns=BASE_CONFIRMED_REQUIRED_COLUMNS,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert result.is_valid
    assert result.missing_columns == ()
    assert result.extra_columns == ()
    assert result.future_columns_present == ()
    assert result.row_count == 1


def test_two_sheet_input_requires_embedded_reference_names(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "duas_abas_incompleto.xlsx"
    create_workbook(
        excel_path,
        base_columns=BASE_CONFIRMED_REQUIRED_COLUMNS,
        include_reference_sheets=False,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert not result.is_valid
    assert (
        result.missing_columns
        == BASE_EMBEDDED_REFERENCE_COLUMNS
    )


def test_two_sheet_input_accepts_embedded_reference_names(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "duas_abas_valido.xlsx"
    create_workbook(
        excel_path,
        base_columns=(
            *BASE_CONFIRMED_REQUIRED_COLUMNS,
            *BASE_EMBEDDED_REFERENCE_COLUMNS,
        ),
        include_reference_sheets=False,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert result.is_valid
    assert result.missing_columns == ()


def test_legacy_profile_accepts_future_columns_as_optional(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "legado_com_futuras.xlsx"
    create_workbook(
        excel_path,
        base_columns=BASE_ALL_COLUMNS,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert result.is_valid
    assert (
        result.future_columns_present
        == BASE_FUTURE_COLUMNS
    )
    assert any(
        issue.code == "BASE-INFO-001"
        for issue in result.information
    )


def test_missing_confirmed_column_blocks_processing(
    tmp_path: Path,
) -> None:
    columns = tuple(
        column
        for column in BASE_CONFIRMED_REQUIRED_COLUMNS
        if column != "idEvento"
    )
    excel_path = tmp_path / "sem_id_evento.xlsx"
    create_workbook(
        excel_path,
        base_columns=columns,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2025-06"),
    )

    assert not result.is_valid
    assert result.missing_columns == ("idEvento",)
    assert any(
        issue.code == "BASE-EST-002"
        and issue.column_name == "idEvento"
        for issue in result.blocking_errors
    )


def test_future_profile_requires_future_columns(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "futuro_incompleto.xlsx"
    create_workbook(
        excel_path,
        base_columns=BASE_CONFIRMED_REQUIRED_COLUMNS,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-12"),
    )

    assert not result.is_valid
    assert (
        result.missing_columns
        == BASE_FUTURE_COLUMNS
    )


def test_future_profile_accepts_all_columns(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "futuro_completo.xlsx"
    create_workbook(
        excel_path,
        base_columns=BASE_ALL_COLUMNS,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-12"),
    )

    assert result.is_valid
    assert result.missing_columns == ()


def test_additional_column_generates_warning_only(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "coluna_extra.xlsx"
    create_workbook(
        excel_path,
        base_columns=(
            *BASE_CONFIRMED_REQUIRED_COLUMNS,
            "observacaoInterna",
        ),
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert result.is_valid
    assert result.extra_columns == (
        "observacaoInterna",
    )
    assert any(
        issue.code == "BASE-AVISO-001"
        for issue in result.warnings
    )


def test_known_alias_is_not_silently_renamed(
    tmp_path: Path,
) -> None:
    columns = tuple(
        (
            "ligadoRiscoSocioambiental"
            if column == "ligacaoRiscoSocioambiental"
            else column
        )
        for column in BASE_CONFIRMED_REQUIRED_COLUMNS
    )
    excel_path = tmp_path / "alias.xlsx"
    create_workbook(
        excel_path,
        base_columns=columns,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert not result.is_valid
    assert (
        "ligacaoRiscoSocioambiental"
        in result.missing_columns
    )
    assert any(
        issue.code == "BASE-EST-003"
        and issue.column_name
        == "ligadoRiscoSocioambiental"
        and issue.suggested_column_name
        == "ligacaoRiscoSocioambiental"
        for issue in result.warnings
    )


def test_empty_base_blocks_processing(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "base_vazia.xlsx"
    create_workbook(
        excel_path,
        base_columns=BASE_CONFIRMED_REQUIRED_COLUMNS,
        include_base_row=False,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert not result.is_valid
    assert result.row_count == 0
    assert any(
        issue.code == "BASE-EST-001"
        for issue in result.blocking_errors
    )


def test_column_order_does_not_matter(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "ordem_diferente.xlsx"
    create_workbook(
        excel_path,
        base_columns=tuple(
            reversed(BASE_CONFIRMED_REQUIRED_COLUMNS)
        ),
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2024-12"),
    )

    assert result.is_valid
    assert result.missing_columns == ()


def test_source_name_is_metadata_information(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "metadata.xlsx"
    create_workbook(
        excel_path,
        base_columns=BASE_CONFIRMED_REQUIRED_COLUMNS,
    )

    result = validate_base_structure(
        read_excel(excel_path),
        get_profile("2026-06"),
    )

    assert any(
        issue.code == "BASE-INFO-002"
        and issue.column_name == "Source.Name"
        for issue in result.information
    )

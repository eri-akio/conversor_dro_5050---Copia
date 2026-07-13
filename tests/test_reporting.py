"""Testes do relatório XLSX."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from src.builders import build_final_document
from src.domain.reporting import (
    FinalExecutionStatus,
)
from src.mappers import group_base_rows
from src.normalizers.header_normalizer import (
    normalize_header,
)
from src.readers import (
    read_and_normalize_base,
    read_excel,
    read_header,
    read_reference_tables,
)
from src.reporters import collect_execution_report
from src.services import (
    calculate_consolidated_events,
    classify_events,
    generate_reports,
    generate_xml,
    resolve_version,
    reconcile_deferred_rules,
    validate_generated_xml,
)
from src.validators import (
    validate_base_rows,
    validate_event_financials,
    validate_grouped_events,
    validate_post_processing,
    validate_pre_processing,
    validate_reference_tables,
)


SAMPLE_PATH = (
    Path(__file__).parent
    / "fixtures"
    / "DRO_5050_planilha_testes.xlsx"
)


def prepare_report_data(tmp_path: Path):
    output_dir = tmp_path / "output"
    started_at = datetime.now().astimezone()

    excel = read_excel(SAMPLE_PATH)
    header_result = normalize_header(
        read_header(excel)
    )
    assert header_result.header is not None
    header = header_result.header

    profile = resolve_version(header).profile
    assert profile is not None

    normalization = read_and_normalize_base(
        excel,
        profile,
    )
    row_validation = validate_base_rows(
        normalization,
        profile,
    )
    grouping = group_base_rows(
        normalization,
        row_validation,
    )
    event_validation = validate_grouped_events(
        grouping,
        profile,
    )
    reconciliation = reconcile_deferred_rules(
        row_validation=row_validation,
        event_validation=event_validation,
    )
    financial = validate_event_financials(
        grouping,
        profile,
    )
    classification = classify_events(
        grouping=grouping,
        event_validation=event_validation,
        financial_validation=financial,
    )
    consolidated = calculate_consolidated_events(
        data_base=header.data_base,
        grouping=grouping,
        classification=classification,
        financial_validation=financial,
    )
    references = validate_reference_tables(
        read_reference_tables(excel),
        grouping,
    )
    pre = validate_pre_processing(
        header=header,
        profile=profile,
        row_validation=row_validation,
        grouping=grouping,
        event_validation=event_validation,
        references=references,
    )
    post = validate_post_processing(
        header=header,
        profile=profile,
        grouping=grouping,
        row_validation=row_validation,
        financial_validation=financial,
        consolidated_events=consolidated.events,
    )
    build = build_final_document(
        header=header,
        profile=profile,
        row_validation=row_validation,
        grouping=grouping,
        event_validation=event_validation,
        financial_validation=financial,
        references=references,
        pre_processing_validation=pre,
        post_processing_validation=post,
        consolidated_events=consolidated.events,
        individualized_event_ids=(
            classification.individualized_event_ids
        ),
    )
    xml = generate_xml(
        build,
        output_dir=output_dir,
    )
    xsd = validate_generated_xml(
        xml,
        profile,
        data_base=header.data_base,
    )

    return collect_execution_report(
        started_at=started_at,
        finished_at=datetime.now().astimezone(),
        input_path=SAMPLE_PATH,
        header=header,
        profile=profile,
        normalization=normalization,
        classification=classification,
        consolidated=consolidated,
        row_validation=row_validation,
        grouping=grouping,
        event_validation=event_validation,
        reconciliation=reconciliation,
        financial_validation=financial,
        references=references,
        pre_processing=pre,
        post_processing=post,
        build_result=build,
        xml_result=xml,
        xsd_result=xsd,
    )


def test_collector_sets_not_apt_status(
    tmp_path: Path,
) -> None:
    data = prepare_report_data(tmp_path)

    assert data.final_status == (
        FinalExecutionStatus.NOT_APT
    )
    assert data.data_base == "2026-06"
    assert data.profile_code == "DRO_2025_06"
    assert len(data.pre_rules) == 34
    assert len(data.post_rules) == 26
    assert data.records
    assert any(
        record.rule_code == "XSD-VAL-001"
        for record in data.records
    )
    assert any(
        record.status == "REGRA NÃO EXECUTADA"
        for record in data.records
    )


def test_generate_xlsx_report_only(
    tmp_path: Path,
) -> None:
    data = prepare_report_data(tmp_path)

    result = generate_reports(
        data,
        output_dir=tmp_path / "saida",
    )

    assert result.is_generated
    assert result.artifacts is not None
    assert result.artifacts.xlsx_path.is_file()
    assert list((tmp_path / "saida").glob("*")) == [
        result.artifacts.xlsx_path
    ]
    assert not (tmp_path / "logs").exists()


def test_xlsx_contains_required_sheets_and_columns(
    tmp_path: Path,
) -> None:
    data = prepare_report_data(tmp_path)
    result = generate_reports(
        data,
        output_dir=tmp_path / "saida",
    )
    assert result.artifacts is not None

    workbook = load_workbook(
        result.artifacts.xlsx_path,
        data_only=False,
    )
    assert workbook.sheetnames == [
        "Resumo",
        "Ocorrencias",
    ]

    occurrences = workbook["Ocorrencias"]
    headers = tuple(
        cell.value for cell in occurrences[1]
    )
    assert len(headers) == 11

    assert headers == (
        "Etapa",
        "Linha",
        "idEvento",
        "Coluna",
        "Valor Original",
        "Valor Normalizado",
        "Regra",
        "Descrição da Regra",
        "Origem",
        "Status",
        "Mensagem",
    )

    for removed_header in (
        "Execução",
        "Data/Hora",
        "Arquivo de Entrada",
        "Arquivo XML",
        "Aba",
        "Dependência",
        "Resultado Final",
        "Versão",
        "Gravidade",
        "Sugestão",
        "Escopo",
        "Resultado Definitivo",
    ):
        assert removed_header not in headers

    summary_values = {
        cell.value
        for row in workbook["Resumo"].iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    }
    for removed_label in (
        "Execução",
        "Início",
        "Fim",
        "Duração (segundos)",
        "dataBase",
        "Perfil",
    ):
        assert removed_label not in summary_values


def test_report_files_are_not_overwritten(
    tmp_path: Path,
) -> None:
    data = prepare_report_data(tmp_path)
    output_dir = tmp_path / "saida"

    first = generate_reports(
        data,
        output_dir=output_dir,
    )
    second = generate_reports(
        data,
        output_dir=output_dir,
    )

    assert first.artifacts is not None
    assert second.artifacts is not None
    assert first.artifacts.xlsx_path != (
        second.artifacts.xlsx_path
    )
    assert second.artifacts.xlsx_path.name.endswith(
        "_001.xlsx"
    )
    assert not (tmp_path / "logs").exists()

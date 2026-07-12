"""Testes da leitura e validação inicial da aba Cabecalho."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from src.config import DOCUMENT_CODE
from src.readers.excel_reader import read_excel
from src.readers.header_reader import (
    HeaderReaderError,
    read_header,
    validate_header_initial,
)


def create_workbook(
    destination: Path,
    *,
    include_document_code: bool = True,
    extra_header_columns: tuple[str, ...] = (),
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    base = workbook.create_sheet("Base")
    base.append(["idEvento"])
    base.append(["EVT001"])

    header = workbook.create_sheet("Cabecalho")

    columns = []
    values = []

    if include_document_code:
        columns.append("codigoDocumento")
        values.append("5050")

    columns.extend(
        [
            "dataBase",
            "codigoConglomerado",
            "cnpj",
            "tipoRemessa",
            "opcaoPorProvisaoAcumulada",
        ]
    )
    values.extend(
        [
            "2026-06",
            "C0099999",
            "99999999",
            "I",
            "N",
        ]
    )

    columns.extend(extra_header_columns)
    values.extend(
        f"metadado-{index}"
        for index, _ in enumerate(
            extra_header_columns,
            start=1,
        )
    )

    header.append(columns)
    header.append(values)

    systems = workbook.create_sheet("Sistemas_Origem")
    systems.append(["codigoSistema", "nomeSistema"])
    systems.append(["SIS001", "Sistema"])

    accounts = workbook.create_sheet("Contas_Internas")
    accounts.append(["codigoConta", "nomeConta"])
    accounts.append(["0001", "Conta"])

    workbook.save(destination)
    workbook.close()


def test_read_valid_header(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "cabecalho_valido.xlsx"
    create_workbook(excel_path)

    excel = read_excel(excel_path)
    header = read_header(excel)
    validation = validate_header_initial(header)

    assert header.sheet_name == "Cabecalho"
    assert header.row_number == 2
    assert header.document_code == "5050"
    assert header.get_value("dataBase") == "2026-06"
    assert header.get_value("tipoRemessa") == "I"
    assert header.extra_columns == ()
    assert validation.is_valid
    assert validation.issues == ()


def test_document_code_is_filled_when_column_is_absent(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "sem_codigo_documento.xlsx"
    create_workbook(
        excel_path,
        include_document_code=False,
    )

    excel = read_excel(excel_path)
    header = read_header(excel)
    validation = validate_header_initial(header)

    field = header.get_field("codigoDocumento")

    assert field.raw_value is None
    assert field.resolved_value == DOCUMENT_CODE
    assert field.source == "FIXED_OFFICIAL"
    assert validation.is_valid
    assert [
        issue.code for issue in validation.information
    ] == ["CAB-INFO-001"]


def test_document_code_is_filled_when_cell_is_blank(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "codigo_vazio.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    workbook["Cabecalho"]["A2"] = None
    workbook.save(excel_path)
    workbook.close()

    excel = read_excel(excel_path)
    header = read_header(excel)
    validation = validate_header_initial(header)

    assert header.document_code == DOCUMENT_CODE
    assert validation.is_valid
    assert any(
        issue.code == "CAB-INFO-001"
        for issue in validation.issues
    )


def test_wrong_document_code_blocks_processing(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "codigo_invalido.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    workbook["Cabecalho"]["A2"] = "9999"
    workbook.save(excel_path)
    workbook.close()

    header = read_header(read_excel(excel_path))
    validation = validate_header_initial(header)

    assert not validation.is_valid
    assert [
        issue.code for issue in validation.blocking_errors
    ] == ["CAB-VAL-003"]


def test_missing_required_column_raises_error(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "sem_cnpj.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    header = workbook["Cabecalho"]
    header.delete_cols(4)
    workbook.save(excel_path)
    workbook.close()

    excel = read_excel(excel_path)

    with pytest.raises(HeaderReaderError) as captured:
        read_header(excel)

    assert captured.value.code == "XLSX-CAB-001"
    assert captured.value.details["colunas_ausentes"] == (
        "cnpj",
    )


def test_header_without_data_row_raises_error(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "sem_linha.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    header = workbook["Cabecalho"]
    header.delete_rows(2, header.max_row)
    workbook.save(excel_path)
    workbook.close()

    excel = read_excel(excel_path)

    with pytest.raises(HeaderReaderError) as captured:
        read_header(excel)

    assert captured.value.code == "XLSX-CAB-002"


def test_more_than_one_header_row_raises_error(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "duas_linhas.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    header = workbook["Cabecalho"]
    header.append(
        [
            "5050",
            "2026-12",
            "C0088888",
            "88888888",
            "S",
            "N",
        ]
    )
    workbook.save(excel_path)
    workbook.close()

    excel = read_excel(excel_path)

    with pytest.raises(HeaderReaderError) as captured:
        read_header(excel)

    assert captured.value.code == "XLSX-CAB-003"
    assert captured.value.details["linhas"] == (2, 3)


def test_empty_required_value_is_reported(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "data_base_vazia.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    workbook["Cabecalho"]["B2"] = "   "
    workbook.save(excel_path)
    workbook.close()

    header = read_header(read_excel(excel_path))
    validation = validate_header_initial(header)

    assert not validation.is_valid
    assert any(
        issue.code == "CAB-VAL-001"
        and issue.field_name == "dataBase"
        for issue in validation.blocking_errors
    )


def test_formula_in_header_is_reported(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "formula.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    workbook["Cabecalho"]["B2"] = '="2026-06"'
    workbook.save(excel_path)
    workbook.close()

    header = read_header(read_excel(excel_path))
    validation = validate_header_initial(header)

    assert not validation.is_valid
    assert any(
        issue.code == "CAB-VAL-002"
        and issue.field_name == "dataBase"
        for issue in validation.blocking_errors
    )


def test_extra_columns_are_informational(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "coluna_extra.xlsx"
    create_workbook(
        excel_path,
        extra_header_columns=("observacaoInterna",),
    )

    header = read_header(read_excel(excel_path))
    validation = validate_header_initial(header)

    assert header.extra_columns == (
        "observacaoInterna",
    )
    assert validation.is_valid
    assert any(
        issue.code == "CAB-INFO-002"
        for issue in validation.information
    )


def test_dash_is_not_considered_empty_in_this_stage(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "hifen.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    workbook["Cabecalho"]["D2"] = "-"
    workbook.save(excel_path)
    workbook.close()

    header = read_header(read_excel(excel_path))
    validation = validate_header_initial(header)

    assert validation.is_valid
    assert header.get_value("cnpj") == "-"

"""Testes da leitura e normalização das linhas da Base."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import to_excel

from src.config import BASE_ALL_COLUMNS
from src.domain.normalization import NormalizationStatus
from src.readers import (
    read_and_normalize_base,
    read_excel,
)
from src.services.version_resolver import resolve_version


def valid_row_values() -> dict[str, object]:
    return {
        "Source.Name": " origem.xlsx ",
        "idEvento": "EVT0001",
        "categoriaNivel1": "1 - Fraudes internas",
        "categoriaNivel2": "11 - Atividade não autorizada",
        "tipoAvaliacao": "I - Individual",
        "unidadeNegocio": "1 - Varejo",
        "dataDescoberta": "30/06/2025",
        "dataOcorrencia": "2025-06-28",
        "totalPerdaEfetiva": "1.427,98",
        "totalProvisao": "0,00",
        "totalRecuperado": "-100,00",
        "valorTotalRisco": "2.000,00",
        "naturezaContingencia": "NA - Não se aplica",
        "codSistemaOrigem": "SISTTI001",
        "codigoEventoOrigem": "ORIG000001",
        "descricaoEvento": (
            "  Evento   sintético para teste.  "
        ),
        "riscoAssociado": "NA",
        "ligacaoRiscoSocioambiental": "N",
        "ligadoRiscoCibernetico": "S - Sim",
        "negocioDescontinuado": "N",
        "idBacen": "Z1234567 - Banco Exemplo",
        "probabilidadePerda": "PR - Provável",
        "valorRisco": "2.000,00",
        "dataContabilizacao": "30/06/2025",
        "contaBalAnaliticoDebito": (
            "810000000000000000000001"
        ),
        "contaBalAnaliticoCredito": (
            "410000000000000000000001"
        ),
        "contaCosifDebito": "10000007",
        "contaCosifCredito": "20000006",
        "valorPerdaEfetiva": "1.427,98",
        "valorProvisao": "0,00",
        "valorRecuperacao": "-100,00",
        "fonteRecuperacao": "S - Seguro",
        "idEventoAgregador": None,
        "dataExclusao": None,
        "motivoExclusao": None,
    }


def create_workbook(
    destination: Path,
    *,
    rows: list[dict[str, object]],
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    base = workbook.create_sheet("Base")
    base.append(list(BASE_ALL_COLUMNS))

    for row_values in rows:
        base.append(
            [
                row_values.get(column_name)
                for column_name in BASE_ALL_COLUMNS
            ]
        )

    header = workbook.create_sheet("Cabecalho")
    header.append(
        [
            "codigoDocumento",
            "dataBase",
            "codigoConglomerado",
            "cnpj",
            "tipoRemessa",
            "opcaoPorProvisaoAcumulada",
        ]
    )
    header.append(
        [
            "5050",
            "2026-06",
            "C0099999",
            "99999999",
            "I",
            "N",
        ]
    )

    systems = workbook.create_sheet("Sistemas_Origem")
    systems.append(["codigoSistema", "nomeSistema"])
    systems.append(["SISTTI001", "Sistema"])

    accounts = workbook.create_sheet("Contas_Internas")
    accounts.append(["codigoConta", "nomeConta"])
    accounts.append(
        [
            "810000000000000000000001",
            "Conta",
        ]
    )

    workbook.save(destination)
    workbook.close()


def inject_cached_formula_result(
    workbook_path: Path,
    *,
    coordinate: str,
    cached_value: str,
    cell_type: str | None = None,
) -> None:
    """Inclui no XLSX um resultado salvo sem calcular a fórmula."""

    temporary_path = workbook_path.with_name(
        workbook_path.stem + "_cached.xlsx"
    )
    namespace = {
        "main": (
            "http://schemas.openxmlformats.org/"
            "spreadsheetml/2006/main"
        )
    }

    with ZipFile(workbook_path, "r") as source, ZipFile(
        temporary_path,
        "w",
        compression=ZIP_DEFLATED,
    ) as destination:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                root = ElementTree.fromstring(content)
                cell = root.find(
                    f".//main:c[@r='{coordinate}']",
                    namespace,
                )
                assert cell is not None
                value_node = cell.find("main:v", namespace)
                if value_node is None:
                    value_node = ElementTree.SubElement(
                        cell,
                        f"{{{namespace['main']}}}v",
                    )
                value_node.text = cached_value
                if cell_type is None:
                    cell.attrib.pop("t", None)
                else:
                    cell.set("t", cell_type)
                content = ElementTree.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                )
            destination.writestr(item, content)

    temporary_path.replace(workbook_path)


def profile_for(data_base: str):
    selection = resolve_version(data_base)
    assert selection.profile is not None
    return selection.profile


def test_valid_row_is_fully_normalized(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "base_valida.xlsx"
    create_workbook(
        excel_path,
        rows=[valid_row_values()],
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )

    assert result.is_valid
    assert result.row_count == 1
    assert result.valid_row_count == 1
    assert result.invalid_row_count == 0

    row = result.rows[0]

    assert row.id_evento == "EVT0001"
    assert (
        row.get_serialized_value("dataOcorrencia")
        == "2025-06-28"
    )
    assert (
        row.get_serialized_value("totalPerdaEfetiva")
        == "1427.98"
    )
    assert (
        row.get_serialized_value("categoriaNivel1")
        == "1"
    )
    assert (
        row.get_field(
            "categoriaNivel1"
        ).result.extracted_description
        == "Fraudes internas"
    )
    assert (
        row.get_serialized_value("idBacen")
        == "Z1234567"
    )
    assert (
        row.get_serialized_value("descricaoEvento")
        == "Evento sintético para teste."
    )


def test_future_fields_are_not_applicable_to_legacy_profile(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["idEventoAgregador"] = "AGREG0001"
    values["dataExclusao"] = "30/06/2025"
    values["motivoExclusao"] = "1 - Exemplo"

    excel_path = tmp_path / "campos_futuros.xlsx"
    create_workbook(
        excel_path,
        rows=[values],
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )
    row = result.rows[0]

    assert row.is_valid
    assert not row.get_field(
        "idEventoAgregador"
    ).applicable
    assert not row.get_field(
        "dataExclusao"
    ).applicable
    assert not row.get_field(
        "motivoExclusao"
    ).applicable
    assert sum(
        issue.code == "BASE-LINHA-INFO-001"
        for issue in row.issues
    ) == 3


def test_future_profile_normalizes_future_fields_without_inventing_domain(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["tipoAvaliacao"] = "IE"
    values["naturezaContingencia"] = "OUT"
    values["idEventoAgregador"] = "AGREG0001"
    values["dataExclusao"] = "31/12/2026"
    values["motivoExclusao"] = "8 - Outro motivo"

    excel_path = tmp_path / "perfil_futuro.xlsx"
    create_workbook(
        excel_path,
        rows=[values],
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-12"),
    )
    row = result.rows[0]

    assert row.is_valid
    assert row.get_serialized_value(
        "tipoAvaliacao"
    ) == "IE"
    assert row.get_serialized_value(
        "naturezaContingencia"
    ) == "OUT"
    assert row.get_serialized_value(
        "motivoExclusao"
    ) == "8"
    assert any(
        issue.code == "BASE-REGRA-NE-001"
        for issue in row.issues
    )


def test_invalid_values_make_only_that_row_invalid(
    tmp_path: Path,
) -> None:
    first = valid_row_values()
    second = valid_row_values()
    second["idEvento"] = "EVT-0002"
    second["totalPerdaEfetiva"] = "1.222"

    excel_path = tmp_path / "linha_invalida.xlsx"
    create_workbook(
        excel_path,
        rows=[first, second],
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )

    assert not result.is_valid
    assert result.valid_row_count == 1
    assert result.invalid_row_count == 1

    assert {
        issue.code
        for issue in result.rows[1].issues
        if issue.blocks_row
    } == {
        "ID-EVENTO-FMT-001",
        "DEC-AMB-001",
    }


def test_absent_value_is_preserved_for_later_rules(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["dataDescoberta"] = None
    values["probabilidadePerda"] = None
    values["valorRisco"] = None

    excel_path = tmp_path / "ausencias.xlsx"
    create_workbook(
        excel_path,
        rows=[values],
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )
    row = result.rows[0]

    assert row.is_valid
    assert row.get_field(
        "dataDescoberta"
    ).status == NormalizationStatus.ABSENT
    assert row.get_field(
        "probabilidadePerda"
    ).status == NormalizationStatus.ABSENT


def test_formula_in_applicable_field_is_invalid(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "formula.xlsx"
    create_workbook(
        excel_path,
        rows=[valid_row_values()],
    )

    workbook = load_workbook(excel_path)
    header_index = {
        cell.value: cell.column
        for cell in workbook["Base"][1]
    }
    workbook["Base"].cell(
        row=2,
        column=header_index["valorPerdaEfetiva"],
        value="=100+50",
    )
    workbook.save(excel_path)
    workbook.close()

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )

    assert not result.is_valid
    assert any(
        issue.code == "BASE-NORM-FORMULA-SEM-RESULTADO-001"
        for issue in result.blocking_issues
    )


def test_monetary_formula_with_cached_result_is_accepted(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "formula_monetaria.xlsx"
    create_workbook(excel_path, rows=[valid_row_values()])
    workbook = load_workbook(excel_path)
    headers = {
        cell.value: cell.column for cell in workbook["Base"][1]
    }
    target = workbook["Base"].cell(
        row=2,
        column=headers["valorPerdaEfetiva"],
        value="=1000+427.98",
    )
    coordinate = target.coordinate
    workbook.save(excel_path)
    workbook.close()
    inject_cached_formula_result(
        excel_path,
        coordinate=coordinate,
        cached_value="1427.98",
    )

    excel = read_excel(excel_path)
    cell = excel.get_sheet("Base").rows[0].get_cell(
        "valorPerdaEfetiva"
    )
    result = read_and_normalize_base(
        excel,
        profile_for("2026-06"),
    )

    assert cell.formula == "=1000+427.98"
    assert cell.cached_value == 1427.98
    assert result.rows[0].get_value("valorPerdaEfetiva") == (
        Decimal("1427.98")
    )
    assert any(
        issue.code == "BASE-NORM-FORMULA-AVISO-001"
        for issue in result.issues
    )


def test_date_formula_with_cached_result_is_accepted(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "formula_data.xlsx"
    create_workbook(excel_path, rows=[valid_row_values()])
    workbook = load_workbook(excel_path)
    headers = {
        cell.value: cell.column for cell in workbook["Base"][1]
    }
    target = workbook["Base"].cell(
        row=2,
        column=headers["dataDescoberta"],
        value="=DATE(2025,6,30)",
    )
    target.number_format = "dd/mm/yyyy"
    coordinate = target.coordinate
    workbook.save(excel_path)
    workbook.close()
    inject_cached_formula_result(
        excel_path,
        coordinate=coordinate,
        cached_value=str(to_excel(date(2025, 6, 30))),
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )

    assert result.rows[0].get_value("dataDescoberta") == date(
        2025,
        6,
        30,
    )


def test_formula_is_prohibited_in_identifier_even_with_cache(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "formula_identificador.xlsx"
    create_workbook(excel_path, rows=[valid_row_values()])
    workbook = load_workbook(excel_path)
    target = workbook["Base"]["B2"]
    assert target.value == "EVT0001"
    target.value = '=CONCAT("EVT","0001")'
    workbook.save(excel_path)
    workbook.close()
    inject_cached_formula_result(
        excel_path,
        coordinate="B2",
        cached_value="EVT0001",
        cell_type="str",
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )

    assert any(
        issue.code == "BASE-NORM-FORMULA-ID-001"
        for issue in result.blocking_issues
    )


def test_formula_with_invalid_cached_result_is_rejected(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "formula_resultado_invalido.xlsx"
    create_workbook(excel_path, rows=[valid_row_values()])
    workbook = load_workbook(excel_path)
    headers = {
        cell.value: cell.column for cell in workbook["Base"][1]
    }
    target = workbook["Base"].cell(
        row=2,
        column=headers["valorPerdaEfetiva"],
        value='="ABC"',
    )
    coordinate = target.coordinate
    workbook.save(excel_path)
    workbook.close()
    inject_cached_formula_result(
        excel_path,
        coordinate=coordinate,
        cached_value="ABC",
        cell_type="str",
    )

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )

    issue = next(
        issue for issue in result.blocking_issues
        if issue.column_name == "valorPerdaEfetiva"
    )
    assert issue.code == "DEC-FMT-001"
    assert issue.original_value == '="ABC"'
    assert issue.normalized_value == "ABC"


def test_non_trivial_money_formats_create_audit_warnings(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["valorPerdaEfetiva"] = "R$ 1.427,98"
    values["valorRecuperacao"] = "(100,00)"
    excel_path = tmp_path / "formatos_monetarios.xlsx"
    create_workbook(excel_path, rows=[values])

    result = read_and_normalize_base(
        read_excel(excel_path),
        profile_for("2026-06"),
    )
    row = result.rows[0]

    assert row.get_value("valorPerdaEfetiva") == Decimal("1427.98")
    assert row.get_value("valorRecuperacao") == Decimal("-100.00")
    assert {
        issue.rule_code for issue in result.issues
        if issue.code == "BASE-NORM-MONETARIO-AVISO-001"
    } == {
        "REMOCAO_SIMBOLO_BRL",
        "CONVERSAO_PARENTESES_CONTABEIS",
    }


def test_cosif_length_depends_on_profile(
    tmp_path: Path,
) -> None:
    values = valid_row_values()
    values["contaCosifDebito"] = "8199510104"

    excel_path = tmp_path / "cosif_10.xlsx"
    create_workbook(
        excel_path,
        rows=[values],
    )
    excel = read_excel(excel_path)

    old_result = read_and_normalize_base(
        excel,
        profile_for("2024-12"),
    )
    current_result = read_and_normalize_base(
        excel,
        profile_for("2026-06"),
    )

    assert not old_result.is_valid
    assert current_result.is_valid


def test_sample_workbook_normalizes_all_30_rows() -> None:
    sample_path = (
        Path(__file__).parent
        / "fixtures"
        / "DRO_5050_planilha_testes.xlsx"
    )

    result = read_and_normalize_base(
        read_excel(sample_path),
        profile_for("2026-06"),
    )

    assert result.row_count == 30
    assert result.valid_row_count == 30
    assert result.invalid_row_count == 0

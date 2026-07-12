"""Testes do leitor principal do Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import pytest

from src.config import REQUIRED_SHEETS
from src.readers.excel_reader import (
    ExcelReaderError,
    ExcelWorkbookReader,
    read_excel,
)


def create_valid_workbook(
    destination: Path,
    *,
    include_additional_sheet: bool = False,
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    base = workbook.create_sheet("Base")
    base.append(["idEvento", "valorPerdaEfetiva"])
    base.append(["EVT001", 1000])
    base.append([None, None])
    base.append(["EVT002", "=SUM(B2,500)"])

    header = workbook.create_sheet("Cabecalho")
    header.append(
        [
            "codigoDocumento",
            "dataBase",
            "codigoConglomerado",
            "cnpj",
            "tipoRemessa",
            "opcaoPorProvisaoAcumulada",
        ]
    )
    header.append(
        [
            "5050",
            "2026-06",
            "C0099999",
            "99999999",
            "I",
            "N",
        ]
    )

    systems = workbook.create_sheet("Sistemas_Origem")
    systems.append(["codigoSistema", "nomeSistema"])
    systems.append(["SIS001", "Sistema de teste"])

    accounts = workbook.create_sheet("Contas_Internas")
    accounts.append(["codigoConta", "nomeConta"])
    accounts.append(["000000000000000000000001", "Conta de teste"])
    accounts["A2"].number_format = "@"

    if include_additional_sheet:
        extra = workbook.create_sheet("Leia_me")
        extra.append(["Aba auxiliar"])

    workbook.save(destination)
    workbook.close()


def test_read_valid_workbook_and_preserve_rows(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "entrada.xlsx"
    create_valid_workbook(excel_path)

    result = read_excel(excel_path)

    assert result.source_path == excel_path.resolve()
    assert tuple(result.sheets) == REQUIRED_SHEETS
    assert result.additional_sheet_names == ()

    base = result.get_sheet("Base")
    assert base.headers == (
        "idEvento",
        "valorPerdaEfetiva",
    )
    assert base.row_count == 2
    assert base.ignored_empty_rows == 1

    first_row = base.rows[0]
    second_row = base.rows[1]

    assert first_row.row_number == 2
    assert first_row.get_value("idEvento") == "EVT001"
    assert first_row.get_value("valorPerdaEfetiva") == 1000

    assert second_row.row_number == 4
    assert second_row.get_value("idEvento") == "EVT002"
    assert (
        second_row.get_value("valorPerdaEfetiva")
        == "=SUM(B2,500)"
    )
    assert (
        second_row
        .get_cell("valorPerdaEfetiva")
        .is_formula
    )

    assert result.total_formulas == 1
    assert result.total_rows == 5


def test_additional_sheets_are_reported_but_not_loaded(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "entrada_com_extra.xlsx"
    create_valid_workbook(
        excel_path,
        include_additional_sheet=True,
    )

    result = ExcelWorkbookReader().read(excel_path)

    assert result.additional_sheet_names == ("Leia_me",)
    assert "Leia_me" not in result.sheets


def test_auxiliary_consolidated_sheet_is_not_loaded_as_source(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "entrada_com_consolidados.xlsx"
    create_valid_workbook(excel_path)

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    consolidated = workbook.create_sheet("Eventos_Consolidados")
    consolidated.append(
        [
            "categoriaNivel1Consol",
            "numEventosTotalConsol",
            "numEventosSemestreConsol",
            "perdaEfetivaTotalConsol",
            "perdaEfetivaSemestreConsol",
            "provisaoTotalConsol",
            "provisaoSemestreConsol",
        ]
    )
    consolidated.append([1, 10, 3, 8000, 2100, 0, 0])
    workbook.save(excel_path)
    workbook.close()

    result = read_excel(excel_path)

    assert "Eventos_Consolidados" not in result.sheets
    assert result.additional_sheet_names == ("Eventos_Consolidados",)


def test_cell_metadata_is_preserved(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "metadados.xlsx"
    create_valid_workbook(excel_path)

    result = read_excel(excel_path)
    account_row = result.get_sheet("Contas_Internas").rows[0]
    account_cell = account_row.get_cell("codigoConta")

    assert account_cell.coordinate == "A2"
    assert account_cell.value == "000000000000000000000001"
    assert account_cell.number_format == "@"


def test_missing_required_sheet_raises_structural_error(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "sem_contas.xlsx"
    create_valid_workbook(excel_path)

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    del workbook["Contas_Internas"]
    workbook.save(excel_path)
    workbook.close()

    with pytest.raises(ExcelReaderError) as captured:
        read_excel(excel_path)

    assert captured.value.code == "XLSX-EST-001"
    assert (
        captured.value.details["abas_ausentes"]
        == ("Contas_Internas",)
    )


def test_duplicate_headers_raise_error(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "cabecalho_duplicado.xlsx"
    create_valid_workbook(excel_path)

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    base = workbook["Base"]
    base["B1"] = "idEvento"
    workbook.save(excel_path)
    workbook.close()

    with pytest.raises(ExcelReaderError) as captured:
        read_excel(excel_path)

    assert captured.value.code == "XLSX-EST-004"


def test_empty_header_name_raises_error(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "cabecalho_vazio.xlsx"
    create_valid_workbook(excel_path)

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path)
    base = workbook["Base"]
    base["A1"] = None
    workbook.save(excel_path)
    workbook.close()

    with pytest.raises(ExcelReaderError) as captured:
        read_excel(excel_path)

    assert captured.value.code == "XLSX-EST-005"


@pytest.mark.parametrize(
    ("filename", "expected_code"),
    [
        ("entrada.xls", "XLSX-READ-002"),
        ("entrada.csv", "XLSX-READ-002"),
        ("inexistente.xlsx", "XLSX-READ-001"),
        ("~$entrada.xlsx", "XLSX-READ-006"),
    ],
)
def test_invalid_input_paths_raise_known_errors(
    tmp_path: Path,
    filename: str,
    expected_code: str,
) -> None:
    path = tmp_path / filename

    with pytest.raises(ExcelReaderError) as captured:
        read_excel(path)

    assert captured.value.code == expected_code


def test_corrupted_xlsx_raises_known_error(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "corrompido.xlsx"
    excel_path.write_text(
        "Isto não é um pacote XLSX.",
        encoding="utf-8",
    )

    with pytest.raises(ExcelReaderError) as captured:
        read_excel(excel_path)

    assert captured.value.code == "XLSX-READ-004"

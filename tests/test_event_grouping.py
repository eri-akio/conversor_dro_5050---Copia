"""Testes do agrupamento e das regras no nível do evento."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.config import BASE_ALL_COLUMNS
from src.domain.base_row_validation import RuleExecutionStatus
from src.mappers import group_base_rows
from src.readers import read_and_normalize_base, read_excel
from src.services.version_resolver import resolve_version
from src.validators import (
    validate_base_rows,
    validate_grouped_events,
)


def base_values() -> dict[str, object]:
    return {
        'Source.Name': 'base.xlsx',
        'idEvento': 'EVT0001',
        'categoriaNivel1': '3',
        'categoriaNivel2': '31',
        'tipoAvaliacao': 'I',
        'unidadeNegocio': '1',
        'dataDescoberta': '2025-06-10',
        'dataOcorrencia': '2025-06-08',
        'totalPerdaEfetiva': '1500,00',
        'totalProvisao': '500,00',
        'totalRecuperado': '0,00',
        'valorTotalRisco': '10500,00',
        'naturezaContingencia': 'TRA',
        'codSistemaOrigem': 'SIST0001',
        'codigoEventoOrigem': 'ORIG0001',
        'descricaoEvento': 'Evento de teste.',
        'riscoAssociado': 'NA',
        'ligacaoRiscoSocioambiental': 'N',
        'ligadoRiscoCibernetico': 'N',
        'negocioDescontinuado': 'N',
        'idBacen': 'Z1234567',
        'probabilidadePerda': 'PR',
        'valorRisco': '10000,00',
        'dataContabilizacao': '2025-06-12',
        'contaBalAnaliticoDebito': '810000000000000000000001',
        'contaBalAnaliticoCredito': '410000000000000000000001',
        'contaCosifDebito': '10000007',
        'contaCosifCredito': '20000006',
        'valorPerdaEfetiva': '1500,00',
        'valorProvisao': '500,00',
        'valorRecuperacao': '0,00',
        'fonteRecuperacao': 'NA',
        'idEventoAgregador': None,
        'dataExclusao': None,
        'motivoExclusao': None,
    }


def create_workbook(
    destination: Path,
    rows: list[dict[str, object]],
    *,
    data_base: str = '2026-06',
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    base = workbook.create_sheet('Base')
    base.append(list(BASE_ALL_COLUMNS))
    for values in rows:
        base.append([
            values.get(column_name)
            for column_name in BASE_ALL_COLUMNS
        ])

    header = workbook.create_sheet('Cabecalho')
    header.append([
        'codigoDocumento',
        'dataBase',
        'codigoConglomerado',
        'cnpj',
        'tipoRemessa',
        'opcaoPorProvisaoAcumulada',
    ])
    header.append([
        '5050',
        data_base,
        'C0099999',
        '99999999',
        'I',
        'N',
    ])

    systems = workbook.create_sheet('Sistemas_Origem')
    systems.append(['codigoSistema', 'nomeSistema'])
    systems.append(['SIST0001', 'Sistema'])

    accounts = workbook.create_sheet('Contas_Internas')
    accounts.append(['codigoConta', 'nomeConta'])
    accounts.append([
        '810000000000000000000001',
        'Débito',
    ])
    accounts.append([
        '410000000000000000000001',
        'Crédito',
    ])

    workbook.save(destination)
    workbook.close()


def process(
    tmp_path: Path,
    rows: list[dict[str, object]],
    *,
    data_base: str = '2026-06',
):
    path = tmp_path / 'entrada.xlsx'
    create_workbook(path, rows, data_base=data_base)

    profile = resolve_version(data_base).profile
    assert profile is not None
    normalization = read_and_normalize_base(
        read_excel(path),
        profile,
    )
    row_validation = validate_base_rows(
        normalization,
        profile,
    )
    grouping = group_base_rows(
        normalization,
        row_validation,
    )
    event_validation = validate_grouped_events(
        grouping,
        profile,
    )
    return grouping, event_validation


def result_for(event_validation, event_id: str, code: str):
    event = next(
        item
        for item in event_validation.events
        if item.id_evento == event_id
    )
    matches = [
        item
        for item in event.rule_results
        if item.code == code
    ]
    assert len(matches) == 1
    return matches[0]


def test_two_rows_create_one_event_and_preserve_accountings(
    tmp_path: Path,
) -> None:
    first = base_values()
    second = base_values()
    second['probabilidadePerda'] = None
    second['valorRisco'] = None
    second['dataContabilizacao'] = '2025-06-20'
    second['valorPerdaEfetiva'] = '0,00'
    second['valorProvisao'] = '0,00'

    grouping, validation = process(tmp_path, [first, second])

    assert grouping.event_count == 1
    event = grouping.events[0]
    assert event.row_numbers == (2, 3)
    assert len(event.accountings) == 2
    assert len(event.probabilities) == 1
    assert event.probabilities[0].source_rows == (2,)
    assert grouping.is_valid
    assert validation.is_valid


def test_conflicting_event_attribute_is_not_chosen(
    tmp_path: Path,
) -> None:
    first = base_values()
    second = base_values()
    second['categoriaNivel1'] = '4'
    second['categoriaNivel2'] = '41'

    grouping, validation = process(tmp_path, [first, second])

    event = grouping.events[0]
    assert not grouping.is_valid
    assert event.get_field('categoriaNivel1').has_conflict
    assert event.get_value('categoriaNivel1') is None
    failure = next(
        item
        for item in event.grouping_results
        if item.code == 'MAP-EVT-001'
    )
    assert failure.status == RuleExecutionStatus.FAILED
    assert not validation.is_valid


def test_duplicate_probability_same_value_is_deduplicated(
    tmp_path: Path,
) -> None:
    first = base_values()
    second = base_values()
    second['dataContabilizacao'] = '2025-06-20'
    second['valorPerdaEfetiva'] = '0,00'
    second['valorProvisao'] = '0,00'

    grouping, validation = process(tmp_path, [first, second])

    event = grouping.events[0]
    assert len(event.probabilities) == 1
    assert event.probabilities[0].source_rows == (2, 3)
    assert event.probabilities[0].value_risk is not None
    assert grouping.is_valid
    assert validation.is_valid


def test_duplicate_probability_with_different_values_fails(
    tmp_path: Path,
) -> None:
    first = base_values()
    second = base_values()
    second['valorRisco'] = '9000,00'
    second['dataContabilizacao'] = '2025-06-20'
    second['valorPerdaEfetiva'] = '0,00'
    second['valorProvisao'] = '0,00'

    grouping, validation = process(tmp_path, [first, second])

    event = grouping.events[0]
    assert event.probabilities[0].has_conflict
    assert not grouping.is_valid
    assert any(
        item.code == 'MAP-PROB-001'
        and item.status == RuleExecutionStatus.FAILED
        for item in event.grouping_results
    )
    assert not validation.is_valid


def test_total_risk_composition_is_validated(
    tmp_path: Path,
) -> None:
    valid_grouping, valid_validation = process(
        tmp_path,
        [base_values()],
    )
    assert valid_grouping.is_valid
    assert result_for(
        valid_validation,
        'EVT0001',
        'DRO001311',
    ).status == RuleExecutionStatus.PASSED

    invalid = base_values()
    invalid['valorTotalRisco'] = '11000,00'
    _, invalid_validation = process(
        tmp_path,
        [invalid],
    )
    assert result_for(
        invalid_validation,
        'EVT0001',
        'DRO001311',
    ).status == RuleExecutionStatus.FAILED


def test_individual_event_requires_probability(
    tmp_path: Path,
) -> None:
    values = base_values()
    values['probabilidadePerda'] = None
    values['valorRisco'] = None
    values['valorTotalRisco'] = None

    _, validation = process(tmp_path, [values])

    assert result_for(
        validation,
        'EVT0001',
        'DRO001312',
    ).status == RuleExecutionStatus.FAILED


def test_positive_risk_sum_is_required(
    tmp_path: Path,
) -> None:
    values = base_values()
    values['valorRisco'] = '0,00'
    values['valorTotalRisco'] = '500,00'

    _, validation = process(tmp_path, [values])

    assert result_for(
        validation,
        'EVT0001',
        'DRO001314',
    ).status == RuleExecutionStatus.FAILED


def test_risk_only_event_cannot_have_accounting_block(
    tmp_path: Path,
) -> None:
    values = base_values()
    values['totalPerdaEfetiva'] = '0,00'
    values['totalProvisao'] = '0,00'
    values['valorTotalRisco'] = '10000,00'
    values['valorPerdaEfetiva'] = '0,00'
    values['valorProvisao'] = '0,00'
    values['valorRecuperacao'] = '0,00'

    _, validation = process(tmp_path, [values])

    assert result_for(
        validation,
        'EVT0001',
        'DRO001452',
    ).status == RuleExecutionStatus.FAILED


def test_sample_workbook_groups_30_rows_into_15_events() -> None:
    path = (
        Path(__file__).parent
        / "fixtures"
        / "DRO_5050_planilha_testes.xlsx"
    )
    profile = resolve_version('2026-06').profile
    assert profile is not None
    normalization = read_and_normalize_base(
        read_excel(path),
        profile,
    )
    row_validation = validate_base_rows(
        normalization,
        profile,
    )
    grouping = group_base_rows(
        normalization,
        row_validation,
    )
    validation = validate_grouped_events(
        grouping,
        profile,
    )

    assert grouping.event_count == 15
    assert grouping.valid_event_count == 15
    assert sum(
        len(event.accountings)
        for event in grouping.events
    ) == 30
    assert sum(
        len(event.probabilities)
        for event in grouping.events
    ) == 3
    assert grouping.is_valid
    assert validation.is_valid
    assert validation.is_fully_verified

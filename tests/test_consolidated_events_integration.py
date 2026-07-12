"""Integração da classificação da Base até XML, XSD e relatório."""

from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

from src.domain.conversion import ConversionStage
from src.domain.base_row_validation import RuleExecutionStatus
from src.domain.reporting import FinalExecutionStatus
from src.domain.xsd_validation import XsdValidationStatus
from src.services import convert_excel


FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "DRO_5050_planilha_produtiva.xlsx"
)


class ConsolidatedEventsIntegrationTests(unittest.TestCase):
    def test_current_workbook_has_only_individualized_events(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            result = convert_excel(FIXTURE, output_dir=directory)
            excel = result.output(ConversionStage.READ_EXCEL)
            classification = result.output(ConversionStage.CLASSIFY_EVENTS)
            calculation = result.output(
                ConversionStage.CALCULATE_CONSOLIDATED
            )
            build = result.output(ConversionStage.BUILD_DOCUMENT)

            self.assertEqual(
                tuple(excel.sheets),
                ("Base", "Cabecalho", "Sistemas_Origem", "Contas_Internas"),
            )
            self.assertEqual(len(classification.individualized_event_ids), 15)
            self.assertEqual(classification.consolidated_event_ids, ())
            self.assertEqual(classification.unresolved_event_ids, ())
            self.assertEqual(calculation.events, ())
            self.assertEqual(
                result.status,
                FinalExecutionStatus.NOT_APT,
            )
            self.assertNotIn(
                "MAP-CONS-001",
                {issue.code for issue in build.issues},
            )
            self.assertIn(
                "DOC-CONS-001",
                {issue.code for issue in build.issues},
            )

            tree = etree.parse(str(result.artifacts.xml_path))
            parent = tree.find("./eventosConsolidados")
            self.assertIsNotNone(parent)
            self.assertEqual(len(parent), 0)

    def test_synthetic_candidate_reaches_xsd_without_duplication(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "massa_mista.xlsx"
            shutil.copyfile(FIXTURE, source)
            self._make_last_event_consolidated(source)

            result = convert_excel(source, output_dir=directory)
            classification = result.output(ConversionStage.CLASSIFY_EVENTS)
            calculation = result.output(
                ConversionStage.CALCULATE_CONSOLIDATED
            )
            build = result.output(ConversionStage.BUILD_DOCUMENT)
            xsd = result.output(ConversionStage.VALIDATE_XSD)
            post = result.output(ConversionStage.POST_PROCESSING)

            self.assertEqual(len(classification.individualized_event_ids), 14)
            self.assertEqual(classification.consolidated_event_ids, ("ORLD0015",))
            self.assertEqual(calculation.event_count, 1)
            self.assertEqual(build.document.individualized_event_count, 14)
            self.assertEqual(build.document.consolidated_event_count, 1)
            self.assertEqual(xsd.status, XsdValidationStatus.VALID)
            for code in (
                "DRO000001",
                "DRO000002",
                "DRO000018",
                "DRO000019",
            ):
                rule = next(
                    item for item in post.rule_results if item.code == code
                )
                self.assertNotEqual(
                    rule.status,
                    RuleExecutionStatus.NOT_EXECUTED,
                )

            tree = etree.parse(str(result.artifacts.xml_path))
            individual_ids = {
                node.get("idEvento")
                for node in tree.findall("./eventosIndividualizados/evento")
            }
            consolidated = tree.findall(
                "./eventosConsolidados/eventoConsolidado"
            )
            self.assertNotIn("ORLD0015", individual_ids)
            self.assertEqual(len(consolidated), 1)
            self.assertEqual(
                consolidated[0].get("numEventosTotalConsol"),
                "1",
            )

    @staticmethod
    def _make_last_event_consolidated(path: Path) -> None:
        workbook = load_workbook(path)
        sheet = workbook["Base"]
        headers = {
            cell.value: cell.column for cell in sheet[1] if cell.value
        }
        rows = [
            row
            for row in range(2, sheet.max_row + 1)
            if sheet.cell(row, headers["idEvento"]).value == "ORLD0015"
        ]
        for index, row in enumerate(rows):
            sheet.cell(row, headers["totalPerdaEfetiva"]).value = 999.99
            sheet.cell(row, headers["totalProvisao"]).value = 0
            sheet.cell(row, headers["totalRecuperado"]).value = 0
            sheet.cell(row, headers["valorTotalRisco"]).value = None
            sheet.cell(row, headers["probabilidadePerda"]).value = None
            sheet.cell(row, headers["valorRisco"]).value = None
            sheet.cell(row, headers["valorPerdaEfetiva"]).value = (
                999.99 if index == 0 else 0
            )
            sheet.cell(row, headers["valorProvisao"]).value = 0
            sheet.cell(row, headers["valorRecuperacao"]).value = 0
        workbook.save(path)
        workbook.close()


if __name__ == "__main__":
    unittest.main()

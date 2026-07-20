"""Testes da normalização e validação do cabeçalho."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
import pytest

from src.readers.excel_reader import read_excel
from src.readers.header_reader import read_header
from src.normalizers.header_normalizer import normalize_header


HEADER_COLUMNS = [
    "codigoDocumento",
    "dataBase",
    "codigoConglomerado",
    "cnpj",
    "tipoRemessa",
    "opcaoPorProvisaoAcumulada",
]


def create_workbook(
    destination: Path,
    header_values: list[object] | None = None,
    *,
    include_document_code: bool = True,
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    base = workbook.create_sheet("Base")
    base.append(["idEvento"])
    base.append(["EVT001"])

    header = workbook.create_sheet("Cabecalho")

    columns = list(HEADER_COLUMNS)
    values = header_values or [
        "5050",
        "2026-06",
        "C0099999",
        "99999999",
        "I",
        "N",
    ]

    if not include_document_code:
        columns = columns[1:]
        values = values[1:]

    header.append(columns)
    header.append(values)

    systems = workbook.create_sheet("Sistemas_Origem")
    systems.append(["codigoSistema", "nomeSistema"])
    systems.append(["SIS001", "Sistema"])

    accounts = workbook.create_sheet("Contas_Internas")
    accounts.append(["codigoConta", "nomeConta"])
    accounts.append(["0001", "Conta"])

    workbook.save(destination)
    workbook.close()


def normalize_from_workbook(
    excel_path: Path,
):
    excel_result = read_excel(excel_path)
    raw_header = read_header(excel_result)
    return normalize_header(raw_header)


def test_valid_header_creates_domain_model(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "valido.xlsx"
    create_workbook(excel_path)

    result = normalize_from_workbook(excel_path)

    assert result.is_valid
    assert result.header is not None
    assert result.header.codigo_documento == "5050"
    assert result.header.data_base == "2026-06"
    assert result.header.codigo_conglomerado == "C0099999"
    assert result.header.cnpj == "99999999"
    assert result.header.tipo_remessa == "I"
    assert (
        result.header.opcao_por_provisao_acumulada
        == "N"
    )
    assert result.header.as_xml_attributes() == {
        "codigoDocumento": "5050",
        "dataBase": "2026-06",
        "codigoConglomerado": "C0099999",
        "cnpj": "99999999",
        "tipoRemessa": "I",
        "opcaoPorProvisaoAcumulada": "N",
    }


def test_fixed_document_code_is_preserved(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "sem_codigo.xlsx"
    create_workbook(
        excel_path,
        include_document_code=False,
    )

    result = normalize_from_workbook(excel_path)

    assert result.is_valid
    assert result.header is not None
    assert result.header.codigo_documento == "5050"

    transformation = next(
        item
        for item in result.transformations
        if item.field_name == "codigoDocumento"
    )
    assert transformation.original_value is None
    assert transformation.normalized_value == "5050"
    assert transformation.changed


@pytest.mark.parametrize(
    ("raw_value", "expected"),
    [
        ("2026-06", "2026-06"),
        ("06/2026", "2026-06"),
        ("30/06/2026", "2026-06"),
        ("2026-06-30", "2026-06"),
        (
            "2026-06-30 10:20:30",
            "2026-06",
        ),
        (date(2026, 6, 30), "2026-06"),
        (
            datetime(2026, 12, 31, 23, 59),
            "2026-12",
        ),
    ],
)
def test_supported_data_base_formats(
    tmp_path: Path,
    raw_value: object,
    expected: str,
) -> None:
    excel_path = tmp_path / "data_base.xlsx"
    create_workbook(
        excel_path,
        [
            "5050",
            raw_value,
            "C0099999",
            "99999999",
            "I",
            "N",
        ],
    )

    result = normalize_from_workbook(excel_path)

    assert result.is_valid
    assert result.header is not None
    assert result.header.data_base == expected


@pytest.mark.parametrize(
    ("raw_value", "error_code"),
    [
        ("2026-01", "CAB-DATA-002"),
        ("2020-06", "CAB-DATA-003"),
        ("2026-13", "CAB-DATA-001"),
        ("31/02/2026", "CAB-DATA-001"),
        ("texto", "CAB-DATA-001"),
        (45000, "CAB-DATA-001"),
    ],
)
def test_invalid_data_base_values(
    tmp_path: Path,
    raw_value: object,
    error_code: str,
) -> None:
    excel_path = tmp_path / "data_base_invalida.xlsx"
    create_workbook(
        excel_path,
        [
            "5050",
            raw_value,
            "C0099999",
            "99999999",
            "I",
            "N",
        ],
    )

    result = normalize_from_workbook(excel_path)

    assert not result.is_valid
    assert result.header is None
    assert any(
        issue.code == error_code
        for issue in result.blocking_errors
    )


def test_cnpj_punctuation_is_removed(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "cnpj_formatado.xlsx"
    create_workbook(
        excel_path,
        [
            "5050",
            "2026-06",
            "C0099999",
            "12.345.678",
            " i ",
            " n ",
        ],
    )

    result = normalize_from_workbook(excel_path)

    assert result.is_valid
    assert result.header is not None
    assert result.header.cnpj == "12345678"
    assert result.header.tipo_remessa == "I"
    assert (
        result.header.opcao_por_provisao_acumulada
        == "N"
    )

    changed_fields = {
        item.field_name
        for item in result.transformations
        if item.changed
    }
    assert "cnpj" in changed_fields
    assert "tipoRemessa" in changed_fields
    assert "opcaoPorProvisaoAcumulada" in changed_fields


def test_lowercase_conglomerate_code_is_normalized(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "conglomerado_minusculo.xlsx"
    create_workbook(
        excel_path,
        [
            "5050",
            "2026-06",
            " c0099999 ",
            "99999999",
            "I",
            "N",
        ],
    )

    result = normalize_from_workbook(excel_path)

    assert result.is_valid
    assert result.header is not None
    assert result.header.codigo_conglomerado == "C0099999"

    transformation = next(
        item
        for item in result.transformations
        if item.field_name == "codigoConglomerado"
    )
    assert transformation.changed
    assert transformation.normalized_value == "C0099999"


@pytest.mark.parametrize(
    "raw_value",
    [
        "39.151.658/0001-30",
        "39151658000130",
    ],
)
def test_full_cnpj_is_normalized_to_eight_digit_root(
    tmp_path: Path,
    raw_value: str,
) -> None:
    excel_path = tmp_path / "cnpj_completo.xlsx"
    create_workbook(
        excel_path,
        [
            "5050",
            "2026-06",
            "C0099999",
            raw_value,
            "I",
            "N",
        ],
    )

    result = normalize_from_workbook(excel_path)

    assert result.is_valid
    assert result.header is not None
    assert result.header.cnpj == "39151658"
    assert not any(
        issue.code == "CAB-CNPJ-001"
        for issue in result.issues
    )


@pytest.mark.parametrize(
    ("field_index", "raw_value", "error_code"),
    [
        (0, "9999", "CAB-DOC-001"),
        (2, "C123", "CAB-CONG-001"),
        (3, "12.345.678/0001-9", "CAB-CNPJ-001"),
        (3, "1234567", "CAB-CNPJ-001"),
        (4, "X", "CAB-REM-001"),
        (5, "T", "CAB-PROV-001"),
        (4, "NA", "CAB-REM-001"),
    ],
)
def test_invalid_header_domains(
    tmp_path: Path,
    field_index: int,
    raw_value: object,
    error_code: str,
) -> None:
    values: list[object] = [
        "5050",
        "2026-06",
        "C0099999",
        "99999999",
        "I",
        "N",
    ]
    values[field_index] = raw_value

    excel_path = tmp_path / f"invalido_{field_index}.xlsx"
    create_workbook(excel_path, values)

    result = normalize_from_workbook(excel_path)

    assert not result.is_valid
    assert any(
        issue.code == error_code
        for issue in result.blocking_errors
    )


@pytest.mark.parametrize(
    ("field_index", "raw_value"),
    [
        (1, "-"),
        (2, "*"),
        (3, "NULL"),
        (4, "N/A"),
    ],
)
def test_null_candidates_block_header(
    tmp_path: Path,
    field_index: int,
    raw_value: object,
) -> None:
    values: list[object] = [
        "5050",
        "2026-06",
        "C0099999",
        "99999999",
        "I",
        "N",
    ]
    values[field_index] = raw_value

    excel_path = tmp_path / f"nulo_{field_index}.xlsx"
    create_workbook(excel_path, values)

    result = normalize_from_workbook(excel_path)

    assert not result.is_valid
    assert any(
        issue.code == "CAB-NULO-001"
        for issue in result.blocking_errors
    )


def test_formula_is_rejected_by_normalizer(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "formula.xlsx"
    create_workbook(excel_path)

    workbook = load_workbook(excel_path)
    workbook["Cabecalho"]["B2"] = '="2026-06"'
    workbook.save(excel_path)
    workbook.close()

    result = normalize_from_workbook(excel_path)

    assert not result.is_valid
    assert any(
        issue.code == "CAB-NORM-002"
        for issue in result.blocking_errors
    )

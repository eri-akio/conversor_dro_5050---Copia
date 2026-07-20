"""Regressão auditável da cadeia XLSX -> XML/XSD -> relatório."""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from defusedxml import ElementTree
from openpyxl import load_workbook
import pytest

from src.domain.conversion import ConversionStage
from src.domain.event_classification import EventDestination
from src.domain.reporting import (
    DependentValidationStatus,
    FinalExecutionStatus,
    GeneralValidationStatus,
    LocalValidationStatus,
    XsdValidationSummaryStatus,
)
from src.domain.xsd_validation import XsdValidationStatus
from src.services import convert_excel

from .workbook_factory import create_workbook, make_row


def _audited_workbook(
    path: Path,
    *,
    embedded_references: bool = False,
) -> Path:
    consolidated_one_rows = tuple(
        make_row(
            event_id="CONS0001",
            category="1",
            occurrence_date=date(2026, 3, 15),
            total_loss=700,
            accounting_loss=value,
        )
        for value in (200, 200, 200, 100)
    )
    rows = (
        make_row(
            event_id="IND-0001",
            category="4",
            overrides={
                "probabilidadePerda": "PO",
                "codigoEventoOrigem": "ORIGIND001",
            },
        ),
        *consolidated_one_rows,
        make_row(
            event_id="CONS0002",
            category="1",
            occurrence_date=date(2025, 12, 15),
            total_loss=450,
            accounting_loss=450,
        ),
        make_row(
            event_id="CONS0003",
            category="2",
            total_loss=900,
            accounting_loss=900,
        ),
        make_row(
            event_id="CONS0004",
            category="4",
            total_loss=650,
            accounting_loss=650,
        ),
        make_row(
            event_id="CONS0005",
            category="7",
            total_loss=800,
            accounting_loss=800,
        ),
    )
    assert len(rows) == 9
    return create_workbook(
        path,
        rows=rows,
        data_base="2026-06",
        embedded_references=embedded_references,
    )


@pytest.mark.parametrize(
    "embedded_references",
    (False, True),
)
def test_audited_xlsx_to_xsd_and_report_chain(
    tmp_path: Path,
    embedded_references: bool,
) -> None:
    input_path = _audited_workbook(
        tmp_path / "entrada_auditada.xlsx",
        embedded_references=embedded_references,
    )
    output_dir = tmp_path / "saida"

    result = convert_excel(input_path, output_dir=output_dir)

    excel = result.output(ConversionStage.READ_EXCEL)
    normalization = result.output(ConversionStage.NORMALIZE_BASE)
    grouping = result.output(ConversionStage.GROUP_EVENTS)
    classification = result.output(ConversionStage.CLASSIFY_EVENTS)
    calculation = result.output(ConversionStage.CALCULATE_CONSOLIDATED)
    build = result.output(ConversionStage.BUILD_DOCUMENT)
    xsd = result.output(ConversionStage.VALIDATE_XSD)
    report_result = result.output(ConversionStage.GENERATE_REPORTS)

    expected_sheets = {"Base", "Cabecalho"}
    if not embedded_references:
        expected_sheets.update({
            "Sistemas_Origem",
            "Contas_Internas",
        })
    assert set(excel.sheets) == expected_sheets
    assert len(normalization.rows) == 9
    assert normalization.rows[0].get_field("totalPerdaEfetiva").is_valid
    assert len(grouping.events) == 6

    destinations = {
        item.event_id: item.destination
        for item in classification.classifications
    }
    assert destinations == {
        "IND0001": EventDestination.INDIVIDUALIZED,
        "CONS0001": EventDestination.CONSOLIDATED,
        "CONS0002": EventDestination.CONSOLIDATED,
        "CONS0003": EventDestination.CONSOLIDATED,
        "CONS0004": EventDestination.CONSOLIDATED,
        "CONS0005": EventDestination.CONSOLIDATED,
    }
    assert not (
        set(classification.individualized_event_ids)
        & set(classification.consolidated_event_ids)
    )

    consolidated = {
        event.category_level_1: event
        for event in calculation.events
    }
    assert set(consolidated) == {"1", "2", "4", "7"}
    expected = {
        "1": (2, Decimal("1150.00"), Decimal("700.00")),
        "2": (1, Decimal("900.00"), Decimal("900.00")),
        "4": (1, Decimal("650.00"), Decimal("650.00")),
        "7": (1, Decimal("800.00"), Decimal("800.00")),
    }
    for category, (count, total, semester) in expected.items():
        event = consolidated[category]
        assert event.total_event_count == count
        assert event.total_loss == total
        assert event.semester_loss == semester
        assert "IND0001" not in event.source_event_ids

    assert build.is_built
    assert build.document is not None
    assert build.document.individualized_event_count == 1
    assert build.document.consolidated_event_count == 4
    assert build.document.source_system_count == 1
    assert build.document.internal_account_count == 2
    assert xsd.status == XsdValidationStatus.VALID
    assert result.status_xsd == XsdValidationSummaryStatus.APPROVED
    assert result.status_local == LocalValidationStatus.APPROVED
    assert result.status_dependencias == DependentValidationStatus.PENDING
    assert result.general_status == GeneralValidationStatus.PENDING
    assert result.status == FinalExecutionStatus.NOT_APT

    id_normalization_records = [
        record
        for record in report_result.data.records
        if record.rule_code == "NORM-ID-EVENTO-001"
        and record.id_evento is None
    ]
    assert len(id_normalization_records) == 1
    id_normalization = id_normalization_records[0]
    assert id_normalization.severity == "INFORMAÇÃO"
    assert id_normalization.status == "NORMALIZADO"
    assert id_normalization.original_value == "IND-0001"
    assert id_normalization.normalized_value == "IND0001"
    assert id_normalization.message == (
        "Separadores permitidos foram removidos do idEvento."
    )

    xml_path = result.artifacts.xml_path
    report_path = result.artifacts.xlsx_path
    assert xml_path is not None and xml_path.parent == output_dir.resolve()
    assert report_path is not None and report_path.parent == output_dir.resolve()
    assert xml_path.is_file() and report_path.is_file()

    root = ElementTree.parse(xml_path).getroot()
    assert root.tag == "documento"
    assert root.get("codigoDocumento") == "5050"
    assert root.get("dataBase") == "2026-06"
    individual_nodes = root.findall("./eventosIndividualizados/evento")
    consolidated_nodes = root.findall(
        "./eventosConsolidados/eventoConsolidado"
    )
    assert [node.get("idEvento") for node in individual_nodes] == [
        "IND0001"
    ]
    assert len(consolidated_nodes) == 4
    assert len(root.findall("./sistemasOrigem/sistema")) == 1
    assert len(root.findall("./contasSubtitulosInternos/conta")) == 2
    xml_text = xml_path.read_text(encoding="utf-8")
    for event_id in ("CONS0001", "CONS0002", "CONS0003", "CONS0004", "CONS0005"):
        assert event_id not in xml_text

    workbook = load_workbook(report_path, data_only=True)
    occurrences = workbook["Ocorrencias"]
    headers = [cell.value for cell in occurrences[1]]
    gravity_column = headers.index("Gravidade") + 1
    status_column = headers.index("Status") + 1
    report_records = report_result.data.records
    assert occurrences.max_row - 1 == len(report_records)
    for row_index, record in enumerate(report_records, start=2):
        assert occurrences.cell(row_index, gravity_column).value == record.severity
        assert occurrences.cell(row_index, status_column).value == record.status

    gravities = {
        occurrences.cell(row, gravity_column).value
        for row in range(2, occurrences.max_row + 1)
    }
    statuses = {
        occurrences.cell(row, status_column).value
        for row in range(2, occurrences.max_row + 1)
    }
    assert gravities.isdisjoint({"APROVADA", "REPROVADA", "PENDENTE"})
    assert statuses.isdisjoint(
        {"ERRO IMPEDITIVO", "ERRO", "AVISO", "INFORMAÇÃO"}
    )

    summary = workbook["Resumo"]
    summary_values = [
        cell.value
        for row in summary.iter_rows()
        for cell in row
    ]
    removed_summary_labels = {
        "Execução",
        "Arquivo de entrada",
        "dataBase",
        "Arquivo XML",
        "Início",
        "Fim",
        "Perfil",
        "XSD selecionado",
    }
    assert removed_summary_labels.isdisjoint(summary_values)
    assert input_path.name not in summary_values
    assert result.execution_id not in summary_values
    assert "2026-06" not in summary_values
    assert str(xml_path) not in summary_values
    assert str(xsd.xsd_path) not in summary_values
    assert not any(
        isinstance(value, str) and value.startswith("O resultado APTO exige")
        for value in summary_values
    )
    assert summary_values.count("PENDENTE") >= 2
    assert FinalExecutionStatus.NOT_APT.value in summary_values
    summary_counts = {
        summary.cell(row, 7).value: summary.cell(row, 8).value
        for row in range(4, 10)
    }
    for severity in ("ERRO IMPEDITIVO", "ERRO", "AVISO", "INFORMAÇÃO"):
        assert summary_counts[severity] == report_result.data.severity_counts.get(
            severity, 0
        )
    assert summary_counts["REGRA NÃO EXECUTADA"] == (
        report_result.data.status_counts.get("REGRA NÃO EXECUTADA", 0)
    )
    assert summary_counts["REPROVADA"] == (
        report_result.data.status_counts.get("REPROVADA", 0)
    )

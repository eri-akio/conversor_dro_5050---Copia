"""Fábrica reproduzível de planilhas para testes de integração."""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from src.config import (
    BASE_ALL_COLUMNS,
    BASE_CREDIT_ACCOUNT_NAME_COLUMN,
    BASE_DEBIT_ACCOUNT_NAME_COLUMN,
    BASE_SOURCE_SYSTEM_NAME_COLUMN,
)


DEBIT_ACCOUNT = "810000000000000000000001"
CREDIT_ACCOUNT = "410000000000000000000001"
SOURCE_SYSTEM = "SISTTI001"

TEXT_COLUMNS = frozenset({
    "idEvento",
    "codSistemaOrigem",
    "codigoEventoOrigem",
    "idBacen",
    "contaBalAnaliticoDebito",
    "contaBalAnaliticoCredito",
    "contaCosifDebito",
    "contaCosifCredito",
    "idEventoAgregador",
})
DATE_COLUMNS = frozenset({
    "dataDescoberta",
    "dataOcorrencia",
    "dataContabilizacao",
    "dataExclusao",
})


def make_row(
    *,
    event_id: str = "IND0001",
    category: str = "1",
    occurrence_date: date = date(2026, 3, 15),
    total_loss: int | float = 2300,
    accounting_loss: int | float = 2300,
    source_system: str = SOURCE_SYSTEM,
    debit_account: str = DEBIT_ACCOUNT,
    credit_account: str = CREDIT_ACCOUNT,
    overrides: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Cria uma linha completa com valores determinísticos."""

    individualized = total_loss >= 1000
    values: dict[str, Any] = {
        "Source.Name": "integracao.xlsx",
        "idEvento": event_id,
        "categoriaNivel1": category,
        "categoriaNivel2": f"{category}1",
        "tipoAvaliacao": "I" if individualized else "NA",
        "unidadeNegocio": category,
        "dataDescoberta": occurrence_date,
        "dataOcorrencia": occurrence_date,
        "totalPerdaEfetiva": total_loss,
        "totalProvisao": 0,
        "totalRecuperado": 0,
        "valorTotalRisco": 10000000 if individualized else None,
        "naturezaContingencia": "TRI" if individualized else "NA",
        "codSistemaOrigem": source_system,
        "codigoEventoOrigem": f"ORIG{event_id[-6:]}",
        "descricaoEvento": "Evento sintético de integração",
        "riscoAssociado": "NA",
        "ligacaoRiscoSocioambiental": "N",
        "ligadoRiscoCibernetico": "N",
        "negocioDescontinuado": "N",
        "idBacen": "Z1234567",
        "probabilidadePerda": "PR" if individualized else None,
        "valorRisco": 10000000 if individualized else None,
        "dataContabilizacao": occurrence_date,
        "contaBalAnaliticoDebito": debit_account,
        "contaBalAnaliticoCredito": credit_account,
        "contaCosifDebito": "10000007",
        "contaCosifCredito": "20000006",
        "valorPerdaEfetiva": accounting_loss,
        "valorProvisao": 0,
        "valorRecuperacao": 0,
        "fonteRecuperacao": "NA",
        "idEventoAgregador": None,
        "dataExclusao": None,
        "motivoExclusao": None,
    }
    if overrides:
        values.update(overrides)
    return values


def create_workbook(
    destination: Path,
    *,
    rows: Iterable[dict[str, Any]],
    data_base: str = "2026-06",
    systems: Iterable[str] = (SOURCE_SYSTEM,),
    accounts: Iterable[str] = (DEBIT_ACCOUNT, CREDIT_ACCOUNT),
    embedded_references: bool = False,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    base = workbook.create_sheet("Base")
    base.append(list(BASE_ALL_COLUMNS))
    for source_values in rows:
        row_values = dict(source_values)
        if embedded_references:
            row_values[BASE_SOURCE_SYSTEM_NAME_COLUMN] = (
                f"Sistema {row_values['codSistemaOrigem']}"
            )
            row_values[BASE_DEBIT_ACCOUNT_NAME_COLUMN] = (
                f"Conta {row_values['contaBalAnaliticoDebito']}"
            )
            row_values[BASE_CREDIT_ACCOUNT_NAME_COLUMN] = (
                f"Conta {row_values['contaBalAnaliticoCredito']}"
            )
        base.append([
            row_values.get(column_name)
            for column_name in BASE_ALL_COLUMNS
        ])

    header_indexes = {
        column_name: index + 1
        for index, column_name in enumerate(BASE_ALL_COLUMNS)
    }
    for row_number in range(2, base.max_row + 1):
        for column_name in TEXT_COLUMNS:
            base.cell(
                row_number,
                header_indexes[column_name],
            ).number_format = "@"
        for column_name in DATE_COLUMNS:
            base.cell(
                row_number,
                header_indexes[column_name],
            ).number_format = "yyyy-mm-dd"

    header = workbook.create_sheet("Cabecalho")
    header.append([
        "codigoDocumento",
        "dataBase",
        "codigoConglomerado",
        "cnpj",
        "tipoRemessa",
        "opcaoPorProvisaoAcumulada",
    ])
    header.append([
        "5050",
        data_base,
        "C0099999",
        "99999999",
        "I",
        "N",
    ])
    for cell in header[2]:
        cell.number_format = "@"

    if not embedded_references:
        systems_sheet = workbook.create_sheet("Sistemas_Origem")
        systems_sheet.append(["codigoSistema", "nomeSistema"])
        for code in systems:
            systems_sheet.append([code, f"Sistema {code}"])
            systems_sheet.cell(
                systems_sheet.max_row,
                1,
            ).number_format = "@"

        accounts_sheet = workbook.create_sheet("Contas_Internas")
        accounts_sheet.append(["codigoConta", "nomeConta"])
        for code in accounts:
            accounts_sheet.append([code, f"Conta {code}"])
            accounts_sheet.cell(
                accounts_sheet.max_row,
                1,
            ).number_format = "@"

    workbook.save(destination)
    workbook.close()
    return destination


def create_valid_workbook(destination: Path) -> Path:
    return create_workbook(destination, rows=(make_row(),))


def create_conflicting_event_workbook(destination: Path) -> Path:
    first = make_row(event_id="EVT0001")
    second = make_row(
        event_id="EVT0001",
        category="2",
    )
    return create_workbook(destination, rows=(first, second))


def create_invalid_money_workbook(destination: Path) -> Path:
    row = make_row(overrides={"valorPerdaEfetiva": "100,123"})
    return create_workbook(destination, rows=(row,))


def create_missing_reference_workbook(destination: Path) -> Path:
    row = make_row(source_system="NAOCAD01")
    return create_workbook(destination, rows=(row,))


def set_formula(
    workbook_path: Path,
    *,
    column_name: str,
    formula: str,
    row_number: int = 2,
    cached_value: str | None = None,
    cached_type: str | None = None,
) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook["Base"]
    headers = {
        cell.value: cell.column for cell in sheet[1]
    }
    target = sheet.cell(
        row_number,
        headers[column_name],
        formula,
    )
    coordinate = target.coordinate
    workbook.save(workbook_path)
    workbook.close()

    if cached_value is not None:
        _inject_cached_formula_result(
            workbook_path,
            coordinate=coordinate,
            cached_value=cached_value,
            cell_type=cached_type,
        )


def _inject_cached_formula_result(
    workbook_path: Path,
    *,
    coordinate: str,
    cached_value: str,
    cell_type: str | None,
) -> None:
    temporary_path = workbook_path.with_name(
        workbook_path.stem + "_cached.xlsx"
    )
    namespace = {
        "main": (
            "http://schemas.openxmlformats.org/"
            "spreadsheetml/2006/main"
        )
    }

    with ZipFile(workbook_path, "r") as source, ZipFile(
        temporary_path,
        "w",
        compression=ZIP_DEFLATED,
    ) as destination:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                root = ElementTree.fromstring(content)
                cell = root.find(
                    f".//main:c[@r='{coordinate}']",
                    namespace,
                )
                assert cell is not None
                value_node = cell.find("main:v", namespace)
                if value_node is None:
                    value_node = ElementTree.SubElement(
                        cell,
                        f"{{{namespace['main']}}}v",
                    )
                value_node.text = cached_value
                if cell_type is None:
                    cell.attrib.pop("t", None)
                else:
                    cell.set("t", cell_type)
                content = ElementTree.tostring(
                    root,
                    encoding="utf-8",
                    xml_declaration=True,
                )
            destination.writestr(item, content)

    temporary_path.replace(workbook_path)
